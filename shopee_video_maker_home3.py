# 蝦皮評論影片製作 - Home Computer Version
# 基於原本 shopee-video-scraper.skill (undetected-chromedriver) 微調
# 新增：Gemini 過濾人臉/開箱、文案標題疊字、BGM

import sys, io, os, re, json, time, random, shutil, tempfile, glob, subprocess, requests, logging
sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8', errors='replace')
sys.stderr = io.TextIOWrapper(sys.stderr.buffer, encoding='utf-8', errors='replace')

import cv2
import warnings; warnings.filterwarnings('ignore')
import numpy as np
import google.generativeai as genai
import openpyxl
from PIL import Image, ImageDraw, ImageFont
from moviepy import VideoFileClip, AudioFileClip, ImageClip, concatenate_videoclips, CompositeVideoClip
import imageio_ffmpeg

logging.basicConfig(level=logging.INFO,
    format='%(asctime)s [%(levelname)s] %(message)s', datefmt='%H:%M:%S')
log = logging.getLogger(__name__)

# ── 路徑設定 ──────────────────────────────────────────────────────────────────
BASE_DIR      = r'C:\Users\user\Desktop\蝦皮自動化工具'
EXCEL_PATH    = os.path.join(BASE_DIR, r'選品Excel\蝦皮關鍵字選品_2026年3-4月.xlsx')
OUTPUT_DIR    = os.path.join(BASE_DIR, '影片輸出')
BGM_DIR       = os.path.join(BASE_DIR, 'BGM音樂')
COOKIES_FILE  = os.path.join(BASE_DIR, 'shopee_cookies.json')
FONT_PATH     = r'C:\Windows\Fonts\msjh.ttc'
FFMPEG_EXE    = imageio_ffmpeg.get_ffmpeg_exe()

# ── Excel 欄位（依實際 Excel 確認）────────────────────────────────────────────
COL_NAME   = 1   # A: 品名
COL_LINK   = 2   # B: 分潤連結
COL_COPY   = 7   # G: 文案
COL_TITLE  = 8   # H: 標題
COL_STATUS = 9   # I: 狀態

# ── 影片參數（沿用 original skill）────────────────────────────────────────────
MIN_CLIPS        = 3    # 最少 3 支才合併
TARGET_DUR       = 18   # 目標秒數
RATINGS_LIMIT    = 10
RATINGS_MAX_PAGE = 30
VIDEO_COLLECT_MAX = 15
VIDEO_W, VIDEO_H = 1080, 1920

# ── Gemini ────────────────────────────────────────────────────────────────────
GEMINI_KEY = 'AIzaSyBfhyW5K3TrNZHs5380tBQ2KjabCmXHtW'
genai.configure(api_key=GEMINI_KEY)
gemini = genai.GenerativeModel(model_name='gemini-1.5-flash')

# ══════════════════════════════════════════════════════════════════════════════
# Cookie helpers（沿用 original skill）
# ══════════════════════════════════════════════════════════════════════════════

def save_cookies(driver):
    with open(COOKIES_FILE, 'w', encoding='utf-8') as f:
        json.dump(driver.get_cookies(), f, ensure_ascii=False)
    log.info('Cookies saved')

def inject_cookies(driver):
    if not os.path.exists(COOKIES_FILE): return False
    with open(COOKIES_FILE, 'r', encoding='utf-8') as f:
        cookies = json.load(f)
    driver.get('https://shopee.tw')
    time.sleep(2)
    for c in cookies:
        try: c.pop('sameSite', None); driver.add_cookie(c)
        except: pass
    log.info('Cookies injected (%d)', len(cookies))
    return True

def is_logged_in(driver):
    return 'buyer/login' not in driver.current_url

def ensure_logged_in(driver):
    if inject_cookies(driver):
        driver.get('https://shopee.tw')
        time.sleep(4)
        if is_logged_in(driver):
            log.info('Logged in via cookies'); return True
        log.warning('Cookies expired')
    driver.get('https://shopee.tw/buyer/login')
    time.sleep(2)
    print('\n' + '='*55)
    print('請在彈出的 Chrome 視窗登入蝦皮，完成後按 Enter')
    print('='*55)
    input()
    if is_logged_in(driver):
        save_cookies(driver); return True
    log.error('Login failed'); return False

# ══════════════════════════════════════════════════════════════════════════════
# ID 解析（沿用 original skill）
# ══════════════════════════════════════════════════════════════════════════════

def extract_ids(url):
    s = str(url or '')
    m = re.search(r'/product/(\d+)/(\d+)', s)
    if m: return int(m.group(1)), int(m.group(2))
    m = re.search(r'-i\.(\d+)\.(\d+)', s)
    if m: return int(m.group(1)), int(m.group(2))
    return None, None

# ══════════════════════════════════════════════════════════════════════════════
# Ratings API（沿用 original skill 的 execute_async_script 方式）
# ══════════════════════════════════════════════════════════════════════════════

def browser_fetch_ratings(driver, shopid, itemid, offset):
    """用 execute_async_script 呼叫 ratings API（繞過反 bot）"""
    for api_ver in ('v2', 'v4'):
        for f_val in (0, 2):
            path = (f'/api/{api_ver}/item/get_ratings'
                    f'?itemid={itemid}&shopid={shopid}'
                    f'&type=0&offset={offset}&limit={RATINGS_LIMIT}&filter={f_val}')
            script = f"""
var cb = arguments[arguments.length - 1];
fetch('{path}', {{credentials: 'include'}})
  .then(r => r.text()).then(t => cb(t)).catch(e => cb(null));
"""
            try:
                raw = driver.execute_async_script(script)
                if not raw: continue
                data = json.loads(raw)
                if data.get('error') and data['error'] != 0: continue
                return data
            except Exception as e:
                log.warning('browser_fetch %s f%d offset %d: %s', api_ver, f_val, offset, e)
    return None

def parse_video_urls(data):
    urls = []
    ratings = (data.get('data') or {}).get('ratings') or []
    for r in ratings:
        for v in (r.get('videos') or []):
            u = v.get('url') or v.get('video_url') or v.get('download_url') or ''
            if u.startswith('http'): urls.append(u)
        for m in (r.get('medias') or []):
            if m.get('type') == 2:
                u = m.get('url') or m.get('video_url') or ''
                if u.startswith('http'): urls.append(u)
        for v in (r.get('review_videos') or []):
            u = v.get('url') or v.get('video_url') or v.get('download_url') or ''
            if u.startswith('http'): urls.append(u)
    return urls, len(ratings)

def get_review_video_urls(driver, product_url):
    """navigate to product page + fetch ratings（沿用 original skill）"""
    shopid, itemid = extract_ids(product_url)
    if shopid is None: return [], 'no_ids'

    # 導到商品頁（undetected-chromedriver 不會被反 bot 擋）
    dest = f'https://shopee.tw/product/{shopid}/{itemid}'
    log.info('Opening: %s', dest)
    driver.get(dest)
    time.sleep(5)

    if 'buyer/login' in driver.current_url:
        return [], 'login_failed'

    driver.set_script_timeout(30)
    video_urls = []
    for page in range(RATINGS_MAX_PAGE):
        offset = page * RATINGS_LIMIT
        data = browser_fetch_ratings(driver, shopid, itemid, offset)
        if data is None: break
        urls, count = parse_video_urls(data)
        video_urls.extend(urls)
        log.info('page %d: %d ratings, %d new vids', page, count, len(urls))
        if count < RATINGS_LIMIT: break
        if len(video_urls) >= VIDEO_COLLECT_MAX: break
        time.sleep(0.4)

    if not video_urls: return [], 'no_videos'
    return list(dict.fromkeys(video_urls))[:VIDEO_COLLECT_MAX], 'ok'

# ══════════════════════════════════════════════════════════════════════════════
# Gemini 影片過濾（沿用 video_filter.py 架構）
# ══════════════════════════════════════════════════════════════════════════════

def gemini_check_frame(img_path):
    """回傳 (has_face, is_unboxing)"""
    try:
        sample = genai.upload_file(path=img_path)
        resp = gemini.generate_content([sample,
            "請回答兩個問題，只能回答「是」或「否」，格式：答案1,答案2\n"
            "問題1：這張圖片中是否有清楚可辨識的人臉？\n"
            "問題2：這張圖片是否為開箱影片（有紙箱正在被拆開）？"])
        text = resp.text.strip().replace('，', ',').replace(' ', '')
        parts = text.split(',')
        return ('是' in parts[0] if parts else False,
                '是' in parts[1] if len(parts) > 1 else False)
    except Exception as e:
        log.warning('[Gemini] %s', e)
        return False, False

def is_valid_video(path):
    try:
        cap = cv2.VideoCapture(path)
        total = int(cap.get(cv2.CAP_PROP_FRAME_COUNT))
        if total == 0: cap.release(); return False
        for pct in (0.15, 0.50, 0.85):
            cap.set(cv2.CAP_PROP_POS_FRAMES, int(total * pct))
            ok, frame = cap.read()
            if not ok: continue
            tmp = path + f'_f{int(pct*100)}.jpg'
            cv2.imwrite(tmp, frame)
            face, unbox = gemini_check_frame(tmp)
            if os.path.exists(tmp): os.remove(tmp)
            if face:  log.info('  X 人臉 @%d%%', int(pct*100)); cap.release(); return False
            if unbox: log.info('  X 開箱 @%d%%', int(pct*100)); cap.release(); return False
        cap.release(); return True
    except Exception as e:
        log.warning('[valid_check] %s', e); return False

# ══════════════════════════════════════════════════════════════════════════════
# 下載
# ══════════════════════════════════════════════════════════════════════════════

def download_video(url, dest):
    try:
        r = requests.get(url, stream=True, timeout=30,
                         headers={'User-Agent': 'Mozilla/5.0', 'Referer': 'https://shopee.tw/'})
        r.raise_for_status()
        with open(dest, 'wb') as f:
            for chunk in r.iter_content(65536): f.write(chunk)
        return os.path.getsize(dest) > 5000
    except Exception as e:
        log.warning('[download] %s', e); return False

# ══════════════════════════════════════════════════════════════════════════════
# 文字疊加工具
# ══════════════════════════════════════════════════════════════════════════════

def make_text_overlay(text, duration, y_pos='bottom', font_size=50, bg_alpha=160):
    try: font = ImageFont.truetype(FONT_PATH, font_size)
    except: font = ImageFont.load_default()
    margin = 40
    dummy = ImageDraw.Draw(Image.new('RGBA', (1, 1)))
    lines, current = [], ''
    for ch in str(text):
        test = current + ch
        if dummy.textlength(test, font=font) > VIDEO_W - margin*2 and current:
            lines.append(current); current = ch
        else: current = test
    if current: lines.append(current)
    if not lines: return None
    line_h = font_size + 12
    box_h = line_h * len(lines) + margin
    img = Image.new('RGBA', (VIDEO_W, box_h), (0, 0, 0, 0))
    draw = ImageDraw.Draw(img)
    draw.rectangle([0, 0, VIDEO_W, box_h], fill=(0, 0, 0, bg_alpha))
    for i, line in enumerate(lines):
        w = draw.textlength(line, font=font)
        x = (VIDEO_W - w) // 2; y = margin//2 + i*line_h
        draw.text((x+2, y+2), line, font=font, fill=(0, 0, 0, 200))
        draw.text((x, y), line, font=font, fill=(255, 255, 255, 255))
    arr = np.array(img)
    pos_y = (VIDEO_H - box_h - 60 if y_pos == 'bottom' else
             60 if y_pos == 'top' else (VIDEO_H - box_h) // 2)
    return (ImageClip(arr, transparent=True)
            .with_duration(duration).with_position(('center', pos_y)))

def get_bgm(duration):
    files = []
    for ext in ('*.mp3', '*.wav', '*.m4a', '*.aac'):
        files += glob.glob(os.path.join(BGM_DIR, ext))
    if not files: return None
    try:
        from moviepy import concatenate_audioclips
        audio = AudioFileClip(random.choice(files))
        if audio.duration < duration:
            loops = int(duration / audio.duration) + 1
            audio = concatenate_audioclips([audio] * loops)
        return audio.subclipped(0, duration).with_volume_scaled(0.4)
    except Exception as e:
        log.warning('[BGM] %s', e); return None

# ══════════════════════════════════════════════════════════════════════════════
# 影片合併 + 後製（沿用 original skill FFmpeg resize + 新增 moviepy 疊字）
# ══════════════════════════════════════════════════════════════════════════════

def get_video_duration(path):
    ffprobe = FFMPEG_EXE.replace('ffmpeg', 'ffprobe')
    if not os.path.exists(ffprobe):
        ffprobe = FFMPEG_EXE  # fallback: use ffmpeg with -i
    try:
        r = subprocess.run(
            [FFMPEG_EXE, '-i', path, '-f', 'null', '-'],
            capture_output=True, text=True, timeout=15)
        m = re.search(r'Duration: (\d+):(\d+):([\d.]+)', r.stderr)
        if m:
            return int(m.group(1))*3600 + int(m.group(2))*60 + float(m.group(3))
    except: pass
    return 0

def produce_video(clip_paths, title, copy_text, output_path):
    """
    1. FFmpeg resize 每支到 1080x1920（沿用 original skill）
    2. moviepy 合併 + 疊標題/文案 + BGM
    """
    with tempfile.TemporaryDirectory() as tmpdir:
        # Step 1: resize（沿用 original skill 的 FFmpeg 指令）
        resized = []
        for i, src in enumerate(clip_paths[:MIN_CLIPS]):
            out = os.path.join(tmpdir, f'clip{i:02d}.mp4')
            cmd = [FFMPEG_EXE, '-y', '-i', src,
                   '-vf', f'scale={VIDEO_W}:{VIDEO_H}:force_original_aspect_ratio=decrease,'
                          f'pad={VIDEO_W}:{VIDEO_H}:(ow-iw)/2:(oh-ih)/2:black',
                   '-c:v', 'libx264', '-an', '-preset', 'fast', '-loglevel', 'error', out]
            subprocess.run(cmd, capture_output=True, timeout=120)
            if os.path.exists(out): resized.append(out)

        if len(resized) < MIN_CLIPS:
            log.warning('resize 後只有 %d 支', len(resized)); return False

        # Step 2: moviepy 合併 + 疊字 + BGM
        try:
            clips = [VideoFileClip(p).without_audio() for p in resized]
            merged = concatenate_videoclips(clips, method='compose')
            if merged.duration > TARGET_DUR:
                merged = merged.subclipped(0, TARGET_DUR)
            dur = merged.duration
            layers = [merged]

            if title and str(title) not in ('標題', 'None', ''):
                ov = make_text_overlay(str(title), min(3.0, dur), 'top', 58)
                if ov: layers.append(ov)

            if copy_text and str(copy_text) not in ('文案', 'None', ''):
                ov = make_text_overlay(str(copy_text), dur, 'bottom', 46)
                if ov: layers.append(ov)

            final = CompositeVideoClip(layers, size=(VIDEO_W, VIDEO_H))
            bgm = get_bgm(dur)
            if bgm:
                final = final.with_audio(bgm)
                log.info('  BGM 已加入')

            final.write_videofile(output_path, codec='libx264', audio_codec='aac',
                fps=30, logger=None,
                ffmpeg_params=['-loglevel', 'error', '-preset', 'fast'])
            for c in clips: c.close()
            merged.close(); final.close()
            return True
        except Exception as e:
            log.error('[produce] %s', e); return False

# ══════════════════════════════════════════════════════════════════════════════
# 主程式
# ══════════════════════════════════════════════════════════════════════════════

def main():
    try:
        import undetected_chromedriver as uc
    except ImportError:
        print('請先執行: pip install undetected-chromedriver'); return

    print('='*55)
    print('蝦皮評論影片製作（undetected-chromedriver 版）')
    print(f'Excel : {EXCEL_PATH}')
    print(f'輸出  : {OUTPUT_DIR}')
    print(f'BGM   : {BGM_DIR}  ← 放 mp3 在這裡就會自動加')
    print('篩選  : 無人臉 / 無開箱 / 最少 3 支合併')
    print('='*55)

    os.makedirs(OUTPUT_DIR, exist_ok=True)

    # 啟動 undetected-chromedriver（自動抓對應 ChromeDriver）
    options = uc.ChromeOptions()
    options.add_argument('--no-sandbox')
    options.add_argument('--disable-dev-shm-usage')
    options.add_argument('--window-size=1280,900')
    options.add_argument('--lang=zh-TW')
    driver = uc.Chrome(options=options)

    try:
        if not ensure_logged_in(driver):
            log.error('無法登入，終止'); return

        wb = openpyxl.load_workbook(EXCEL_PATH)
        ws = wb.active
        total = ws.max_row - 1
        done_count = 0

        for row in ws.iter_rows(min_row=2):
            name        = row[COL_NAME   - 1].value
            link        = row[COL_LINK   - 1].value
            copy_text   = row[COL_COPY   - 1].value
            title       = row[COL_TITLE  - 1].value
            status_cell = row[COL_STATUS - 1]
            row_idx     = row[0].row - 1

            if not name or not link: continue
            if status_cell.value in ('影片完成', '無評論影片'):
                log.info('[skip %d] %s', row_idx, str(name)[:25]); continue

            print(f'\n[{row_idx:3}/{total}] {str(name)[:45]}')
            log.info('標題: %s', str(title)[:30] if title else '（空）')
            log.info('文案: %s', str(copy_text)[:30] if copy_text else '（空）')

            # 抓評論影片 URL（undetected-chromedriver → 不被反 bot 擋）
            video_urls, reason = get_review_video_urls(driver, str(link))
            log.info('評論影片: %d 個 (%s)', len(video_urls), reason)

            if not video_urls:
                status_cell.value = reason; wb.save(EXCEL_PATH); continue

            # 下載 + Gemini 過濾
            tmpdir = tempfile.mkdtemp()
            valid_clips = []
            try:
                for i, url in enumerate(video_urls):
                    if len(valid_clips) >= MIN_CLIPS: break
                    dest = os.path.join(tmpdir, f'clip_{i:02d}.mp4')
                    log.info('[%2d/%d] 下載...', i+1, len(video_urls))
                    if not download_video(url, dest): continue
                    log.info('  Gemini 檢查...')
                    if is_valid_video(dest):
                        valid_clips.append(dest)
                        log.info('  V 通過 (%d/%d)', len(valid_clips), MIN_CLIPS)
                    time.sleep(0.5)

                log.info('有效: %d / 需要: %d', len(valid_clips), MIN_CLIPS)

                if len(valid_clips) >= MIN_CLIPS:
                    safe = re.sub(r'[\\/:*?"<>|]', '', str(name))[:35]
                    out  = os.path.join(OUTPUT_DIR, f'{row_idx:03d}_{safe}.mp4')
                    log.info('後製 → %s', os.path.basename(out))
                    if produce_video(valid_clips, title, copy_text, out):
                        status_cell.value = '影片完成'; done_count += 1
                    else:
                        status_cell.value = '合併失敗'
                else:
                    status_cell.value = f'影片不足({len(valid_clips)}/{MIN_CLIPS})'
            finally:
                shutil.rmtree(tmpdir, ignore_errors=True)

            wb.save(EXCEL_PATH)
            wait = random.uniform(8, 15)
            log.info('等待 %.1fs...', wait)
            time.sleep(wait)

        print(f'\n{"="*55}')
        print(f'完成！成功製作 {done_count} 個商品影片')
        print(f'存放：{OUTPUT_DIR}')

    finally:
        try: driver.quit()
        except: pass

if __name__ == '__main__':
    main()
