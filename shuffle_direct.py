"""
shuffle_direct.py
把 Excel 中「蝦皮直營」的列隨機打散插入非直營列之間（13筆以後）
執行前自動備份原檔
"""
import sys, io, os, random, shutil, copy
sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8', errors='replace')

from dotenv import load_dotenv
load_dotenv(dotenv_path=os.path.join(os.path.dirname(os.path.abspath(__file__)), '.env'))
import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment, Border
from copy import copy as cp

EXCEL_PATH = os.getenv('EXCEL_PATH', r'C:\Users\user\Desktop\蝦皮素材\蝦皮選品_2026年5月new.xlsx')
COL_NAME   = 2  # B 品名
KEEP_ROWS  = 12 # 前幾筆不動

def snapshot_rows(ws, min_row, max_row):
    """把每一列所有欄位的 value + fill 存成 list of list of dict"""
    result = []
    for row in ws.iter_rows(min_row=min_row, max_row=max_row):
        cells = []
        for cell in row:
            fill = None
            if cell.fill and cell.fill.fill_type not in (None, 'none'):
                fill = cp(cell.fill)
            cells.append({'value': cell.value, 'fill': fill})
        result.append(cells)
    return result

def write_rows(ws, rows_data, start_row):
    """把快照寫回 worksheet"""
    for idx, cells in enumerate(rows_data):
        row_num = start_row + idx
        for col_idx, c in enumerate(cells, start=1):
            cell = ws.cell(row=row_num, column=col_idx, value=c['value'])
            if c['fill']:
                cell.fill = c['fill']

def main():
    print(f'Excel: {EXCEL_PATH}')
    if not os.path.exists(EXCEL_PATH):
        print('找不到 Excel！'); return

    # 備份
    backup = EXCEL_PATH.replace('.xlsx', '_backup_shuffle.xlsx')
    shutil.copy2(EXCEL_PATH, backup)
    print(f'已備份至: {backup}')

    wb = openpyxl.load_workbook(EXCEL_PATH)
    ws = wb.active
    total = ws.max_row - 1  # 不含 header

    # ── 先快照所有資料（delete_rows 前）──────────────────────────────────────
    fixed_data = snapshot_rows(ws, min_row=2,            max_row=1+KEEP_ROWS)
    rest_data  = snapshot_rows(ws, min_row=2+KEEP_ROWS,  max_row=1+total)

    print(f'固定不動（1~{KEEP_ROWS}）: {len(fixed_data)} 筆')

    # ── 分成直營 / 非直營 ────────────────────────────────────────────────────
    direct, non_direct = [], []
    for row in rest_data:
        name = str(row[COL_NAME - 1]['value'] or '')
        if '蝦皮直營' in name:
            direct.append(row)
        else:
            non_direct.append(row)

    print(f'第{KEEP_ROWS+1}筆以後 — 蝦皮直營: {len(direct)} 筆 / 非直營: {len(non_direct)} 筆')

    if not direct:
        print('13筆以後沒有蝦皮直營，不需要打散'); return

    # ── 隨機插入：大約每 gap 筆非直營插一筆直營 ──────────────────────────────
    gap = max(1, len(non_direct) // max(len(direct), 1))
    random.shuffle(direct)

    shuffled, di = [], 0
    for i, row in enumerate(non_direct):
        shuffled.append(row)
        if di < len(direct) and (i + 1) % gap == 0:
            shuffled.append(direct[di]); di += 1
    while di < len(direct):
        pos = random.randint(0, len(shuffled))
        shuffled.insert(pos, direct[di]); di += 1

    result = fixed_data + shuffled

    # ── 清除並寫回 ────────────────────────────────────────────────────────────
    ws.delete_rows(2, ws.max_row)
    write_rows(ws, result, start_row=2)

    # 重新編號 A 欄（13筆以後）
    for i in range(len(result)):
        ws.cell(row=2+i, column=1, value=i+1)

    wb.save(EXCEL_PATH)
    print(f'\n完成！1~{KEEP_ROWS} 不動，第{KEEP_ROWS+1}~{len(result)} 已隨機打散')
    print(f'直營平均每 {gap} 筆出現一次')
    print(f'存檔：{EXCEL_PATH}')

if __name__ == '__main__':
    main()
