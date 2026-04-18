# 蝦皮選品影片製作 - Home Computer Version
# 讀取選品Excel → 抓評論影片 → Gemini篩(無人臉/無開箱) → 合併3支 → 加文案字幕 + BGM → 輸出

import sys, io, asyncio, json, re, os, random, requests, time, tempfile, shutil, glob, subprocess
os.environ['PATH'] += os.pathsep + r'C:\ffmpeg\bin'
os.environ['PATH'] += os.pathsep + r'C:\platform-tools'
sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8', errors='replace')
sys.stderr = io.TextIOWrapper(sys.stderr.buffer, encoding='utf-8', errors='replace')

import cv2
import warnings; warnings.filterwarnings('ignore')
import google.generativeai as genai
import openpyxl
from PIL import Image, ImageDraw, ImageFont
from playwright.async_api import async_playwright
from moviepy import VideoFileClip, AudioFileClip, ImageClip, concatenate_videoclips, CompositeVideoClip
import imageio_ffmpeg
import numpy as np

# ── 路徑設定 ──────────────────────────────────────────────────────────────────
BASE_DIR   = r'C:\Users\user\Desktop\蝦皮自動化工具'
EXCEL_PATH = r'C:\Users\user\Desktop\蝦皮素材\蝦皮關鍵字選品_2026年3-4月new.xlsx'
OUTPUT_DIR = r'C:\Users\user\Desktop\蝦皮素材\蝦皮影片輸出'
BGM_DIR    = r'C:\Users\user\Desktop\蝦皮素材\BGM音樂'
FONT_PATH  = r'C:\Windows\Fonts\msjh.ttc'    # 微軟正黑體（支援繁體中文）

# ── 欄位對應（依照實際Excel）─────────────────────────────────────────────────
COL_NAME   = 2   # B: 品名
COL_LINK   = 3   # C: 分潤連結
COL_COPY   = 8   # H: 文案
COL_TITLE  = 9   # I: 標題
COL_STATUS = 10  # J: 狀態

# ── 影片參數 ──────────────────────────────────────────────────────────────────
MIN_CLIPS    = 3    # 最少幾支影片才合併
MAX_TRY      = 15   # 每個商品最多試幾支評論影片
TARGET_SEC   = 18   # 合併後目標秒數
VIDEO_W      = 1080
VIDEO_H      = 1920

# ── Gemini 設定 ───────────────────────────────────────────────────────────────
GEMINI_KEY = 'AIzaSyBfhyW5K3TrNZHs5380tBQ2KjabCmXHtW'
genai.configure(api_key=GEMINI_KEY)
gemini = genai.GenerativeModel(model_name='gemini-1.5-flash')

FFMPEG_EXE = r'C:\ffmpeg\bin\ffmpeg.exe'

# ═══════════════════════════════════════════════════════════════════════════════
# 文字工具
# ═══════════════════════════════════════════════════════════════════════════════

def wrap_text(text, font, max_width, draw):
    """自動換行，回傳 list of lines"""
    if not text: return []
    lines, current = [], ''
    for ch in str(text):
        test = current + ch
        w = draw.textlength(test, font=font)
        if w > max_width and current:
            lines.append(current)
            current = ch
        else:
            current = test
    if current: lines.append(current)
    return lines

def make_text_overlay(text, video_w, video_h, duration,
                      font_size=52, y_pos='bottom', bg_alpha=160):
    """
    用 PIL 渲染文字，回傳 moviepy ImageClip（透明背景疊在影片上）
    y_pos: 'top' | 'bottom' | 'center'
    """
    try:
        font = ImageFont.truetype(FONT_PATH, font_size)
    except:
        font = ImageFont.load_default()

    # 計算文字區域
    margin = 40
    max_w  = video_w - margin * 2
    dummy  = ImageDraw.Draw(Image.new('RGBA', (1, 1)))
    lines  = wrap_text(text, font, max_w, dummy)
    if not lines: return None

    line_h   = font_size + 12
    box_h    = line_h * len(lines) + margin
    box_w    = video_w

    img = Image.new('RGBA', (box_w, box_h), (0, 0, 0, 0))
    draw = ImageDraw.Draw(img)

    # 半透明黑底
    draw.rectangle([0, 0, box_w, box_h], fill=(0, 0, 0, bg_alpha))

    # 文字
    for i, line in enumerate(lines):
        w = draw.textlength(line, font=font)
        x = (box_w - w) // 2
        y = margin // 2 + i * line_h
        # 陰影
        draw.text((x+2, y+2), line, font=font, fill=(0, 0, 0, 200))
        draw.text((x, y), line, font=font, fill=(255, 255, 255, 255))

    arr = np.array(img)

    if y_pos == 'bottom':
        pos_y = video_h - box_h - 60
    elif y_pos == 'top':
        pos_y = 60
    else:
        pos_y = (video_h - box_h) // 2

    clip = (ImageClip(arr, transparent=True)
            .with_duration(duration)
            .with_position(('center', pos_y)))
    return clip


def make_title_card(title, copy_text, duration=2.5):
    """開頭標題卡（黑底白字）"""
    img = Image.new('RGB', (VIDEO_W, VIDEO_H), (10, 10, 10))
    draw = ImageDraw.Draw(img)

    try:
        font_big  = ImageFont.truetype(FONT_PATH, 64)
        font_small = ImageFont.truetype(FONT_PATH, 40)
    except:
        font_big = font_small = ImageFont.load_default()

    # 標題
    if title:
        lines = wrap_text(str(title), font_big, VIDEO_W - 80, draw)
        total_h = len(lines) * 80
        start_y = (VIDEO_H - total_h) // 2 - 60
        for i, line in enumerate(lines):
            w = draw.textlength(line, font=font_big)
            draw.text(((VIDEO_W - w)//2 + 2, start_y + i*80 + 2), line,
                      font=font_big, fill=(0,0,0))
            draw.text(((VIDEO_W - w)//2, start_y + i*80), line,
                      font=font_big, fill=(255,255,255))

    arr = np.array(img)
    return ImageClip(arr).with_duration(duration)


# ═══════════════════════════════════════════════════════════════════════════════
# BGM 工具
# ═══════════════════════════════════════════════════════════════════════════════

def get_bgm(video_duration):
    """
    從 BGM音樂 資料夾隨機選一首，裁切/循環到指定秒數
    沒有音樂就回傳 None
    """
    exts = ['*.mp3', '*.wav', '*.m4a', '*.aac']
    music_files = []
    for ext in exts:
        music_files += glob.glob(os.path.join(BGM_DIR, ext))
    if not music_files:
        return None
    chosen = random.choice(music_files)
    try:
        audio = AudioFileClip(chosen)
        # 如果音樂比影片短，循環
        if audio.duration < video_duration:
            loops = int(video_duration / audio.duration) + 1
            import moviepy
            from moviepy import concatenate_audioclips
            audio = concatenate_audioclips([audio] * loops)
        audio = audio.subclipped(0, video_duration).with_volume_scaled(0.4)
        return audio
    except Exception as e:
        print(f'  [BGM] 載入失敗: {e}')
        return None


# ═══════════════════════════════════════════════════════════════════════════════
# Gemini 影片過濾
# ═══════════════════════════════════════════════════════════════════════════════

def gemini_check_frame(img_path):
    """單張截圖：回傳 (has_face, is_unboxing)"""
    try:
        sample = genai.upload_file(path=img_path)
        resp = gemini.generate_content([
            sample,
            "請回答兩個問題，每個只能回答「是」或「否」，格式：答案1,答案2\n"
            "問題1：這張圖片中是否有清楚可辨識的人臉？\n"
            "問題2：這張圖片是否為開箱影片（有紙箱或包裝盒正在被拆開）？"
        ])
        text = resp.text.strip().replace('，', ',').replace(' ', '')
        parts = text.split(',')
        face     = len(parts) > 0 and '是' in parts[0]
        unboxing = len(parts) > 1 and '是' in parts[1]
        return face, unboxing
    except Exception as e:
        print(f'      [Gemini] {e}')
        return False, False

def is_valid_video(video_path):
    """抽 3 個時間點截圖送 Gemini，任一有人臉或開箱就拒絕"""
    try:
        cap = cv2.VideoCapture(video_path)
        total = int(cap.get(cv2.CAP_PROP_FRAME_COUNT))
        if total == 0:
            cap.release(); return False

        for pct in (0.15, 0.50, 0.85):
            cap.set(cv2.CAP_PROP_POS_FRAMES, int(total * pct))
            ok, frame = cap.read()
            if not ok: continue

            tmp = video_path + f'_f{int(pct*100)}.jpg'
            cv2.imwrite(tmp, frame)
            face, unbox = gemini_check_frame(tmp)
            if os.path.exists(tmp): os.remove(tmp)

            if face:
                print(f' X 人臉@{int(pct*100)}%'); cap.release(); return False
            if unbox:
                print(f' X 開箱@{int(pct*100)}%'); cap.release(); return False

        cap.release()
        return True
    except Exception as e:
        print(f'  [valid_check] {e}')
        return False


# ═══════════════════════════════════════════════════════════════════════════════
# 影片下載
# ═══════════════════════════════════════════════════════════════════════════════

def download_video(url, dest):
    try:
        r = requests.get(url, stream=True, timeout=30,
                         headers={'User-Agent': 'Mozilla/5.0',
                                  'Referer': 'https://shopee.tw/'})
        r.raise_for_status()
        with open(dest, 'wb') as f:
            for chunk in r.iter_content(65536): f.write(chunk)
        return os.path.getsize(dest) > 5000
    except Exception as e:
        print(f'  [download] {e}'); return False


# ═══════════════════════════════════════════════════════════════════════════════
# 影片合併 + 後製（文案疊字 + BGM）
# ═══════════════════════════════════════════════════════════════════════════════

def resize_clip(clip):
    """統一縮放到 1080x1920（直式），不足補黑邊"""
    import subprocess, tempfile as tf
    src  = clip.filename if hasattr(clip, 'filename') else None
    if src is None: return clip

    out = src + '_resized.mp4'
    cmd = [
        FFMPEG_EXE, '-y', '-i', src,
        '-vf', f'scale={VIDEO_W}:{VIDEO_H}:force_original_aspect_ratio=decrease,'
               f'pad={VIDEO_W}:{VIDEO_H}:(ow-iw)/2:(oh-ih)/2:black',
        '-c:v', 'libx264', '-an', '-preset', 'fast', '-loglevel', 'error', out
    ]
    subprocess.run(cmd, capture_output=True, timeout=120)
    if os.path.exists(out):
        return VideoFileClip(out)
    return clip


FONT_PATH_FFMPEG = FONT_PATH.replace('\\', '/').replace(':', '\\:')

def build_drawtext(lines, dur):
    """每句依等分時間顯示，位置在畫面正中央"""
    if not lines:
        return ''
    per = dur / len(lines)
    parts = []
    for i, line in enumerate(lines):
        start = i * per
        end   = start + per
        safe  = (line
                 .replace('\\', '\\\\')
                 .replace("'",  '\u2019')
                 .replace(':',  '\\:')
                 .replace('%',  '\\%'))
        parts.append(
            f"drawtext=fontfile='{FONT_PATH_FFMPEG}'"
            f":text='{safe}'"
            f":fontcolor=white:fontsize=52"
            f":x=(w-text_w)/2:y=(h-text_h)/2"
            f":box=1:boxcolor=black@0.50:boxborderw=14"
            f":shadowcolor=black@0.8:shadowx=2:shadowy=2"
            f":enable='between(t,{start:.3f},{end:.3f})'"
        )
    return ','.join(parts)


def produce_video(clip_paths, title, copy_text, output_path):
    """
    1. 合併 clip_paths（去音訊）→ 暫存檔
    2. FFmpeg drawtext 逐句字幕（畫面正中央，每句輪流出現）
    3. 加 BGM
    4. 輸出 mp4
    """
    import tempfile as tf
    tmp_merged = output_path + '_merged.mp4'
    try:
        # 1. moviepy 合併
        clips = []
        for p in clip_paths[:MIN_CLIPS]:
            c = VideoFileClip(p).without_audio()
            clips.append(c)

        merged = concatenate_videoclips(clips, method='compose')
        if merged.duration > TARGET_SEC:
            merged = merged.subclipped(0, TARGET_SEC)
        dur = merged.duration

        merged.write_videofile(
            tmp_merged, codec='libx264', audio=False, fps=30,
            logger=None, ffmpeg_params=['-loglevel', 'error', '-preset', 'fast']
        )
        for c in clips: c.close()
        merged.close()

        # 2. FFmpeg 逐句字幕 + BGM
        lines = [l.strip() for l in str(copy_text or '').split('。') if l.strip()]
        vf    = build_drawtext(lines, dur) if lines else 'null'

        bgm_files = []
        for ext in ('*.mp3', '*.wav', '*.m4a', '*.aac'):
            bgm_files += glob.glob(os.path.join(BGM_DIR, ext))

        if bgm_files:
            bgm_path = random.choice(bgm_files)
            cmd = [
                FFMPEG_EXE, '-y',
                '-stream_loop', '-1', '-i', tmp_merged,
                '-stream_loop', '-1', '-i', bgm_path,
                '-vf', vf,
                '-filter_complex',
                f'[1:a]volume=0.15,atrim=0:{dur:.3f}[bgm]',
                '-map', '0:v', '-map', '[bgm]',
                '-c:v', 'libx264', '-preset', 'fast',
                '-c:a', 'aac', '-b:a', '128k',
                '-t', f'{dur:.3f}',
                '-loglevel', 'error', output_path
            ]
            print(f'  BGM: 已加入')
        else:
            cmd = [
                FFMPEG_EXE, '-y', '-i', tmp_merged,
                '-vf', vf,
                '-c:v', 'libx264', '-preset', 'fast', '-an',
                '-loglevel', 'error', output_path
            ]
            print(f'  BGM: 無音樂（可放 mp3 到 BGM音樂 資料夾）')

        r = subprocess.run(cmd, capture_output=True, timeout=300)
        if r.returncode != 0:
            print(f'  [ffmpeg] {r.stderr.decode(errors="ignore")[-500:]}')
            return False
        return True
    except Exception as e:
        print(f'  [produce] {e}')
        return False
    finally:
        if os.path.exists(tmp_merged):
            os.remove(tmp_merged)


# ═══════════════════════════════════════════════════════════════════════════════
# Shopee ratings API（CDP fetch）
# ═══════════════════════════════════════════════════════════════════════════════

def extract_ids(url):
    s = str(url or '')
    m = re.search(r'/product/(\d+)/(\d+)', s)
    if m: return m.group(1), m.group(2)
    m = re.search(r'-i\.(\d+)\.(\d+)', s)
    if m: return m.group(1), m.group(2)
    return None, None


async def fetch_review_videos(page, shop_id, item_id):
    video_urls = []
    for offset in range(0, 10 * 30, 10):
        js = f"""
        (async () => {{
          try {{
            const r = await fetch(
              '/api/v2/item/get_ratings?itemid={item_id}&shopid={shop_id}' +
              '&type=0&offset={offset}&limit=10&filter=0',
              {{credentials: 'include'}}
            );
            return await r.json();
          }} catch(e) {{ return {{error: e.toString()}}; }}
        }})()
        """
        data = await page.evaluate(js)
        if not data or data.get('error'): break

        ratings = (data.get('data') or {}).get('ratings') or []
        if not ratings: break

        for rating in ratings:
            for v in (rating.get('videos') or []):
                url = v.get('url') or v.get('video_url') or v.get('download_url') or ''
                if url.startswith('http'): video_urls.append(url)
            for m in (rating.get('medias') or []):
                if m.get('type') == 2:
                    url = m.get('url') or m.get('video_url') or ''
                    if url.startswith('http'): video_urls.append(url)

        if len(ratings) < 10: break
        if len(video_urls) >= MAX_TRY: break
        await asyncio.sleep(0.3)

    return list(dict.fromkeys(video_urls))[:MAX_TRY]


# ═══════════════════════════════════════════════════════════════════════════════
# 處理單一商品
# ═══════════════════════════════════════════════════════════════════════════════

async def process_product(page, row_data, row_idx):
    name, link, copy_text, title = row_data
    shop_id, item_id = extract_ids(link)
    if not shop_id:
        print(f'  X 無法解析 ID')
        return '無法解析ID'

    print(f'  shop={shop_id}  item={item_id}')

    # 確保頁面不在 captcha
    if 'captcha' in page.url or 'verify' in page.url:
        await page.goto('https://shopee.tw', wait_until='domcontentloaded', timeout=15000)
        await asyncio.sleep(2)

    video_urls = await fetch_review_videos(page, shop_id, item_id)
    print(f'  評論影片: {len(video_urls)} 個')
    if not video_urls:
        return '無評論影片'

    tmpdir = os.path.join(OUTPUT_DIR, f'_clips_{row_idx:03d}')
    os.makedirs(tmpdir, exist_ok=True)
    valid_clips = []

    try:
        for i, url in enumerate(video_urls):
            if len(valid_clips) >= MIN_CLIPS: break

            dest = os.path.join(tmpdir, f'clip_{i:02d}.mp4')
            print(f'  [{i+1:2}/{len(video_urls)}] 下載...', end=' ', flush=True)
            if not download_video(url, dest):
                print('失敗'); continue

            print('Gemini...', end=' ', flush=True)
            if is_valid_video(dest):
                valid_clips.append(dest)
                print(f'通過 ({len(valid_clips)}/{MIN_CLIPS})')
            time.sleep(0.5)

        print(f'  有效: {len(valid_clips)} / 需要: {MIN_CLIPS}')

        if len(valid_clips) >= MIN_CLIPS:
            safe_name = re.sub(r'[\\/:*?"<>|]', '', str(name))[:35]
            out_path  = os.path.join(OUTPUT_DIR, f'{row_idx:03d}_{safe_name}.mp4')
            print(f'  後製 → {os.path.basename(out_path)}')
            if produce_video(valid_clips, title, copy_text, out_path):
                return '影片完成'
            else:
                return '合併失敗'
        else:
            return f'影片不足({len(valid_clips)}/{MIN_CLIPS})'

    finally:
        pass  # 原始片段保留在 _clips_XXX 資料夾，後製完成後可手動刪除


# ═══════════════════════════════════════════════════════════════════════════════
# 主程式
# ═══════════════════════════════════════════════════════════════════════════════

async def main():
    print('=' * 60)
    print('蝦皮選品影片製作 v2')
    print(f'Excel : {EXCEL_PATH}')
    print(f'輸出  : {OUTPUT_DIR}')
    print(f'BGM   : {BGM_DIR}（放 mp3 進去就會自動加音樂）')
    print('篩選  : 無人臉 / 無開箱 / 每品最少 3 支合併')
    print('後製  : 標題疊字(上) + 文案疊字(下) + BGM')
    print('=' * 60)

    os.makedirs(OUTPUT_DIR, exist_ok=True)

    wb = openpyxl.load_workbook(EXCEL_PATH)
    ws = wb.active
    total = ws.max_row - 1
    done_count = 0

    async with async_playwright() as p:
        print('\n連接 Chrome (port 9222)...')
        browser = await p.chromium.connect_over_cdp('http://127.0.0.1:9222')
        ctx = browser.contexts[0]

        pg = next((pg for pg in ctx.pages
                   if pg.url.startswith('https://shopee.tw') and 'affiliate' not in pg.url), None)
        if not pg:
            pg = await ctx.new_page()
            await pg.goto('https://shopee.tw', wait_until='domcontentloaded', timeout=20000)
            await asyncio.sleep(3)
        print(f'頁面: {pg.url[:60]}')

        for row in ws.iter_rows(min_row=2):
            name        = row[COL_NAME   - 1].value
            link        = row[COL_LINK   - 1].value
            copy_text   = row[COL_COPY   - 1].value
            title       = row[COL_TITLE  - 1].value
            status_cell = row[COL_STATUS - 1]
            row_idx     = row[0].row - 1

            if not name or not link: continue
            if status_cell.value in ('影片完成', '無評論影片'):
                print(f'[skip {row_idx:3}] {str(name)[:25]}')
                continue

            print(f'\n[{row_idx:3}/{total}] {str(name)[:45]}')
            print(f'  標題: {str(title)[:30] if title else "（空）"}')
            print(f'  文案: {str(copy_text)[:30] if copy_text else "（空）"}')

            status = await process_product(
                pg, (name, link, copy_text, title), row_idx)

            status_cell.value = status
            wb.save(EXCEL_PATH)
            if status == '影片完成': done_count += 1
            await asyncio.sleep(random.uniform(1.5, 3.0))

        pass  # keep browser open

    print(f'\n{"="*60}')
    print(f'完成！成功製作 {done_count} 個商品影片')
    print(f'影片存放：{OUTPUT_DIR}')

asyncio.run(main())
