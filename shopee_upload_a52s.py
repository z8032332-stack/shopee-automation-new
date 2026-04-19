# -*- coding: utf-8 -*-
"""
蝦皮短影音批次上傳腳本
用法：
  python shopee_upload.py                              # 上傳全部待上傳的影片
  python shopee_upload.py --row 2                      # 只上傳第2筆
  python shopee_upload.py --dry-run                    # 跑到發佈前停下
  python shopee_upload.py --row 2 --dry-run            # 第2筆 dry-run
  python shopee_upload.py --phone mi_note2             # 用小米 Note 2
  python shopee_upload.py --phone s25fe                # 用 Samsung S25 FE
  python shopee_upload.py --phone a52s                 # 用 Samsung A52s 5G
  python shopee_upload.py --device 192.168.0.29:5555   # 手動指定裝置

設定檔（config.json，放在同目錄）：
  {
    "excel_path": "C:\\path\\to\\蝦皮分潤前100_整理版.xlsx",
    "video_dir":  "C:\\path\\to\\output_final",
    "adb_path":   "C:\\platform-tools\\adb.exe",
    "screenshot_dir": "C:\\path\\to\\screenshots"
  }
  ※ 沒有 config.json 則使用下方預設值（此電腦）
"""

import os, sys, time, re, argparse, subprocess, json, glob
from dotenv import load_dotenv
load_dotenv(dotenv_path=os.path.join(os.path.dirname(os.path.abspath(__file__)), '.env'))

# 修正 Windows 終端機編碼
if sys.platform == 'win32':
    sys.stdout.reconfigure(encoding='utf-8', errors='replace')
    sys.stderr.reconfigure(encoding='utf-8', errors='replace')

import openpyxl
import uiautomator2 as u2

# ── 預設路徑（從 .env 讀取，沒設定則用此電腦預設值）──
_DEFAULTS = {
    "excel_path":     os.getenv("EXCEL_PATH",     r"D:\Users\user\Downloads\Annie\Claud\蝦皮專案\蝦皮分潤前100_整理版.xlsx"),
    "video_dir":      os.getenv("VIDEO_DIR",      r"D:\Users\user\Desktop\蝦皮影片專案\output_final"),
    "adb_path":       os.getenv("ADB_PATH",       r"D:\platform-tools\adb.exe"),
    "screenshot_dir": os.getenv("SCREENSHOT_DIR", r"D:\Users\user\Desktop\蝦皮影片專案\screenshots"),
}

# ── 讀取 config.json（若存在則覆蓋預設值）──
_cfg_path = os.path.join(os.path.dirname(os.path.abspath(__file__)), "config.json")
_cfg = {}
if os.path.exists(_cfg_path):
    with open(_cfg_path, encoding="utf-8") as _f:
        _cfg = json.load(_f)

EXCEL_PATH     = _cfg.get("excel_path",     _DEFAULTS["excel_path"])
VIDEO_DIR      = _cfg.get("video_dir",      _DEFAULTS["video_dir"])
ADB_PATH       = _cfg.get("adb_path",       _DEFAULTS["adb_path"])
SCREENSHOT_DIR = _cfg.get("screenshot_dir", _DEFAULTS["screenshot_dir"])

DEVICE          = os.getenv("PHONE_IP", "192.168.0.9:5555")  # Samsung Galaxy A52s 5G（預設）
PHONE_VIDEO_DIR = "/sdcard/DCIM/ShopeeUpload"

# ── 手機座標 Profiles ──
PHONE_PROFILES = {
    # Samsung S25 FE, 1080x2340
    "s25fe": {
        "shortvideo_tab": (540, 2150),
        "plus_btn":       (1000, 153),
        "media_lib":      (862, 1849),
        "video_tab":      (540, 307),
        "first_video":    (133, 506),
        "next_btn":       (930, 2000),   # 影片選擇頁「下一步」
        "editor_next":    (970, 2070),   # 編輯頁「下一步」
    },
    # 小米 Note 2, 1080x1920（shortvideo_tab 實測；其餘待校準）
    "mi_note2": {
        "shortvideo_tab": (540, 1897),   # 實測
        "plus_btn":       (1000, 126),   # 待校準
        "media_lib":      (862, 1518),   # 待校準
        "video_tab":      (540, 252),    # 待校準
        "first_video":    (133, 415),    # 待校準
    },
    # Samsung Galaxy A52s 5G, 1080x2400（實測）
    "a52s": {
        "device":         os.getenv("PHONE_IP", "192.168.0.9:5555"),
        "shortvideo_tab": (540, 2205),
        "plus_btn":       (1000, 157),
        "media_lib":      (862, 1899),
        "video_tab":      (539, 310),    # 短影音 tab 實測（UI dump）
        "first_video":    (180, 553),    # 3欄 grid 第一格中心（tab底 Y=373，格高 360px）
        "next_btn":       (956, 2180),   # 影片選擇頁「下一步」（UI dump 實測）
        "editor_next":    (934, 2136),   # 編輯頁「下一步」（實測）
        "add_product":        (540, 1076),  # 發布頁「新增商品」列（估算）
        "all_categories_tab": (540, 230),  # 商品搜尋頁「全部分類」tab（待校準）
        "product_search_box": (700, 130),  # 商品搜尋框
        "hq_btn":             (1157, 179), # 編輯頁右側「畫質優化」圖示
    },
    # 其他手機（自行填入座標後使用 --phone custom）
    "custom": {
        "shortvideo_tab": (540, 0),   # ← 請填入
        "plus_btn":       (0,   0),   # ← 請填入
        "media_lib":      (0,   0),   # ← 請填入
        "video_tab":      (0,   0),   # ← 請填入
        "first_video":    (0,   0),   # ← 請填入
    },
}

COORD = PHONE_PROFILES["a52s"]  # 預設使用 Samsung Galaxy A52s 5G


def adb(cmd):
    """執行 ADB 指令"""
    full = f'"{ADB_PATH}" -s {DEVICE} {cmd}'
    r = subprocess.run(full, shell=True, capture_output=True, timeout=30)
    out = r.stdout.decode('utf-8', errors='replace') if r.stdout else ''
    err = r.stderr.decode('utf-8', errors='replace') if r.stderr else ''
    return out + err


def screenshot(name):
    """截圖並儲存"""
    os.makedirs(SCREENSHOT_DIR, exist_ok=True)
    path = os.path.join(SCREENSHOT_DIR, f"{name}.png")
    subprocess.run(
        f'"{ADB_PATH}" -s {DEVICE} exec-out screencap -p',
        shell=True, stdout=open(path, 'wb'), timeout=15
    )
    return path


_d = None  # uiautomator2 device instance，連線後設定

def tap(x, y, wait=1.5):
    """點擊螢幕（優先用 u2 HTTP API，MIUI 封鎖 adb input 時仍可用）"""
    if _d is not None:
        _d.click(x, y)
    else:
        adb(f"shell input tap {x} {y}")
    time.sleep(wait)


def swipe(x1, y1, x2, y2, duration=300, wait=1):
    """滑動"""
    adb(f"shell input swipe {x1} {y1} {x2} {y2} {duration}")
    time.sleep(wait)


def connect_device():
    """連接 ADB 裝置"""
    if ":" in DEVICE:
        result = adb(f"connect {DEVICE}")
        print(f"  ADB: {result.strip()}")
        time.sleep(1)
    else:
        result = adb("get-state")
        print(f"  ADB (USB): {DEVICE} → {result.strip()}")


def unlock_screen(pin="0000"):
    """喚醒螢幕並解鎖（PIN 輸入）"""
    # 1. 確認螢幕狀態
    state = adb("shell dumpsys power")
    if "Display Power: state=OFF" in state or "mWakefulness=Asleep" in state:
        adb("shell input keyevent KEYCODE_WAKEUP")
        time.sleep(1)

    # 2. 確認是否在鎖定畫面
    win = adb("shell dumpsys window")
    if "mDreamingLockscreen=true" in win or "StatusBar" in win:
        # 向上滑開鎖定畫面
        adb("shell input swipe 540 1800 540 600 300")
        time.sleep(1)
        # 輸入 PIN
        adb(f"shell input text {pin}")
        time.sleep(0.5)
        adb("shell input keyevent KEYCODE_ENTER")
        time.sleep(1)
        print(f"  🔓 螢幕已解鎖")


def push_video(local_path):
    """推送影片到手機（先清空舊檔確保只有一個影片）"""
    # 1. 從 MediaStore 刪除 ShopeeUpload 資料夾的所有記錄
    #    用 relative_path（Android 10+ 可用）
    adb("shell content delete --uri content://media/external/video/media"
        " --where \"relative_path='DCIM/ShopeeUpload/'\"")
    time.sleep(1)

    # 2. 清空資料夾
    adb(f"shell rm -rf {PHONE_VIDEO_DIR}")
    time.sleep(1)
    adb(f"shell mkdir -p {PHONE_VIDEO_DIR}")

    # 3. 推送新影片（固定用 upload_temp.mp4 避免中文空格路徑問題）
    filename = os.path.basename(local_path)
    remote = f"{PHONE_VIDEO_DIR}/upload_temp.mp4"
    print(f"  推送影片: {filename}")
    result = subprocess.run(
        [ADB_PATH, "-s", DEVICE, "push", local_path, remote],
        capture_output=True, timeout=120
    )
    push_out = (result.stdout + result.stderr).decode('utf-8', errors='replace')
    print(f"  push 結果: {push_out.strip()}")

    # 4. touch 更新 mtime（強制 scanner 認定為新檔案）
    adb(f'shell touch "{remote}"')
    time.sleep(1)

    # 5. 確認檔案存在
    check = adb(f'shell ls -la "{remote}"')
    print(f"  檔案確認: {check.strip()}")

    # 6. 強制媒體掃描
    adb(f'shell am broadcast -a android.intent.action.MEDIA_SCANNER_SCAN_FILE -d "file://{remote}"')
    time.sleep(2)

    # 7. 等待媒體掃描完成（固定 5 秒）
    print("  等待媒體掃描...", end='', flush=True)
    time.sleep(5)
    print(" ✓")
    return remote


def cleanup_phone_video(remote_path):
    """上傳完成後刪掉手機上的影片"""
    adb(f'shell rm -f "{remote_path}"')
    adb(f'shell am broadcast -a android.intent.action.MEDIA_SCANNER_SCAN_FILE -d "file://{remote_path}"')


def open_shopee():
    """開啟蝦皮 APP（先強制關閉確保乾淨起始狀態）"""
    print("  開啟蝦皮...")
    adb("shell am force-stop com.shopee.tw")
    time.sleep(2)
    adb("shell monkey -p com.shopee.tw -c android.intent.category.LAUNCHER 1")
    time.sleep(6)


def close_all_apps():
    """關閉所有分頁"""
    print("  關閉所有分頁...")
    if _d:
        _d.press("recent")
    else:
        adb("shell input keyevent KEYCODE_APP_SWITCH")
    time.sleep(2)
    d = _d if _d else u2.connect(DEVICE)
    btn = d(text='關閉全部')
    if btn.exists(timeout=3):
        btn.click()
        time.sleep(2)
    else:
        if _d:
            _d.press("home")
        else:
            adb("shell input keyevent KEYCODE_HOME")
        time.sleep(1)


def find_and_click(d, texts, timeout=5, label="元素"):
    """嘗試多個文字找到並點擊元素，回傳是否成功（優先 textContains）"""
    if isinstance(texts, str):
        texts = [texts]
    for txt in texts:
        # 先用 textContains（處理「下一步(1)」這種情況）
        el = d(textContains=txt)
        if el.exists(timeout=timeout):
            el.click()
            return True
    print(f"  ⚠ 找不到{label}: {texts}")
    return False


def close_popup(d):
    """嘗試關閉短影音 feed 的彈窗"""
    time.sleep(2)
    for _ in range(3):
        if d(text='短影音').exists(timeout=1) or d(text='推薦').exists(timeout=1):
            break
        # 找關閉按鈕（文字元素，不用座標）
        closed = False
        for close_txt in ['稍後', '關閉', '跳過', 'X', '×', '我知道了']:
            el = d(text=close_txt)
            if el.exists(timeout=1):
                el.click()
                time.sleep(1)
                closed = True
                break
        if not closed:
            # 若不在 feed 頁（可能是全頁 promo），按返回鍵
            if not d(text='推薦').exists(timeout=1):
                adb("shell input keyevent KEYCODE_BACK")
                time.sleep(2)
            break
    time.sleep(1)


def navigate_to_create(d):
    """導航到短影音拍攝頁"""
    print("  進入短影音...")
    # 用文字元素點底部 tab（不需要座標）
    if not find_and_click(d, ['直播短影音'], timeout=5, label="直播短影音tab"):
        tap(*COORD["shortvideo_tab"], wait=3)  # fallback
    time.sleep(3)

    # 關閉可能的彈窗
    close_popup(d)

    print("  點 + 按鈕...")
    # 蝦皮短影音拍攝頁的建立按鈕（content-desc 或文字）
    found = False
    for desc in ['拍攝', '建立', '新增', '短片']:
        el = d(description=desc)
        if el.exists(timeout=2):
            el.click()
            found = True
            break
    if not found:
        tap(*COORD["plus_btn"], wait=3)  # fallback 座標

    time.sleep(3)
    screenshot("after_plus")


def select_video_from_gallery(d):
    """從媒體庫選擇最新的影片"""
    print("  點媒體庫...")
    if not find_and_click(d, ['媒體庫', '相簿', '圖庫'], timeout=5, label="媒體庫"):
        tap(*COORD["media_lib"], wait=3)  # fallback
    time.sleep(3)

    screenshot("gallery_page")

    print("  篩選短影音...")
    if not find_and_click(d, ['短影音', '影片'], timeout=3, label="短影音tab"):
        tap(*COORD["video_tab"], wait=2)  # fallback
    time.sleep(4)   # 等相簿穩定
    screenshot("after_videotab")

    print("  選第一個影片...")
    screenshot("before_tap_video")  # 確認相簿還在
    # 用 adb shell input tap（繞過 uiautomator2，避免 Samsung 上的點擊異常）
    x, y = COORD["first_video"]
    adb(f"shell input tap {x} {y}")
    time.sleep(3)
    screenshot("after_select_video")

    print("  下一步（影片預覽）...")
    # 先用文字找（gallery 或 publish page 的元素，timeout 短）
    found = find_and_click(d, ['下一步', '繼續', 'Next'], timeout=3, label="影片預覽下一步")
    if not found:
        el = d(description='下一步')
        if el.exists(timeout=2):
            el.click()
            found = True
    if not found:
        # 若在編輯器且工具列可能隱藏：先 tap 影片中心喚醒工具列，再打座標
        screenshot("no_next_btn")
        print("  → 喚醒工具列後 tap next_btn")
        adb("shell input tap 540 1000")  # 喚醒工具列
        time.sleep(1)
        adb(f"shell input tap {COORD['next_btn'][0]} {COORD['next_btn'][1]}")
        time.sleep(3)
    screenshot("after_next_btn")

    # 處理「上傳畫質建議」彈窗
    time.sleep(2)
    ok_btn = d(text='OK')
    if ok_btn.exists(timeout=4):
        ok_btn.click()
        print("  ✓ 關閉上傳畫質建議")
        time.sleep(3)
        screenshot("after_popup_ok")  # 看 OK 後的狀態
        # OK 後直接點下一步（不重選，避免二次觸發 popup）
        print("  點下一步進入編輯頁...")
        found2 = find_and_click(d, ['下一步', '繼續', 'Next'], timeout=5, label="重試下一步")
        if not found2:
            tap(*COORD["next_btn"], wait=3)
    time.sleep(3)


def skip_editor(d):
    """跳過編輯頁，直接下一步（先開畫質優化）"""
    print("  跳過編輯頁...")
    time.sleep(4)
    screenshot("editor_page")

    # 若已在發布頁（撰寫內文 / 新增商品 / 發佈 任一存在），直接跳過
    if (d(textContains='撰寫內文').exists(timeout=2)
            or d(textContains='新增商品').exists(timeout=2)
            or d(text='發佈').exists(timeout=2)):
        print("  ✓ 已在發布頁，無需再點下一步")
        screenshot("after_editor")
        return

    # ── 開啟畫質優化（編輯頁右側圖示）──
    hq_found = False
    for _txt in ['畫質優化', '畫質', '高畫質']:
        _el = d(textContains=_txt)
        if _el.exists(timeout=2):
            _el.click()
            time.sleep(1)
            print("  ✓ 已開啟畫質優化")
            hq_found = True
            break
    if not hq_found:
        # React Native fallback：右側第三個圖示座標
        _hx, _hy = COORD.get("hq_btn", (1157, 179))
        adb(f"shell input tap {_hx} {_hy}")
        time.sleep(1)
        print(f"  → ADB tap 畫質優化 ({_hx},{_hy})")

    found = find_and_click(d, ['下一步', '繼續', 'Next'], timeout=3, label="編輯頁下一步")
    if not found:
        el = d(description='下一步')
        if el.exists(timeout=2):
            el.click()
            found = True
    if not found:
        screenshot("no_editor_next")
        print("  → 喚醒工具列後 tap editor_next")
        adb("shell input tap 540 1000")  # 喚醒工具列
        time.sleep(1)
        adb(f"shell input tap {COORD['editor_next'][0]} {COORD['editor_next'][1]}")
        time.sleep(3)

    time.sleep(5)
    screenshot("after_editor")


def enter_caption(d, caption_text):
    """輸入文案"""
    print("  輸入文案...")
    # 找到文案輸入框 - 嘗試多種方式
    found = False
    for txt in ['為您的短影音撰寫內文', '撰寫內文']:
        el = d(textContains=txt)
        if el.exists(timeout=3):
            el.click()
            time.sleep(1)
            found = True
            break

    if not found:
        # fallback: 點文案區域座標（發布頁上方）
        tap(540, 300, wait=1)

    # 限制 150 字
    if len(caption_text) > 150:
        caption_text = caption_text[:150]
    # 直接對元素 set_text（不走 clipboard，避免 u2.jar threading 問題）
    for txt in ['為您的短影音撰寫內文', '撰寫內文']:
        el = d(textContains=txt)
        if el.exists(timeout=3):
            el.set_text(caption_text)
            time.sleep(1)
            break
    else:
        # fallback：點欄位後直接 set_text 第一個 EditText
        tap(540, 300, wait=1)
        el = d(className='android.widget.EditText')
        if el.exists(timeout=3):
            el.set_text(caption_text)
            time.sleep(1)

    # 收鍵盤 - 點右上角 OK
    ok = d(text='OK')
    if ok.exists(timeout=3):
        ok.click()
        time.sleep(2)
    else:
        # 點其他空白區域收鍵盤
        tap(540, 600, wait=1)

    screenshot("after_caption")
    print(f"  ✓ 文案已輸入（{len(caption_text)}字）")


def toggle_off_switches(d):
    """關閉允許合拍和允許拼接"""
    print("  關閉合拍/拼接...")

    for label in ['允許他人合拍', '允許他人拼接']:
        try:
            el = d(text=label)
            if not el.exists(timeout=3):
                print(f"    ⚠ 找不到 {label}，跳過")
                continue
            # 優先找同一容器內的 Switch 元素
            switch = el.sibling(className='android.widget.Switch')
            if not switch.exists(timeout=1):
                switch = el.sibling(className='android.widget.ToggleButton')
            if switch.exists(timeout=1):
                info = switch.info
                if info.get('checked', True):  # 已開啟才關
                    switch.click()
                    time.sleep(1.5)
                    print(f"    ✓ 已關閉 {label}")
                else:
                    print(f"    ✓ {label} 已是關閉狀態")
            else:
                # fallback 座標
                bounds = el.info['bounds']
                toggle_y = (bounds['top'] + bounds['bottom']) // 2
                tap(980, toggle_y, wait=1.5)
                print(f"    ✓ 已點 {label} toggle")
        except Exception as _e:
            print(f"    ⚠ {label} toggle 例外: {_e}，跳過")

    screenshot("after_toggle")


def _dump_ui_find_product_btn(d):
    """UI dump 找「新增商品」可點擊父容器，回傳 (cx, cy) 或 None"""
    import xml.etree.ElementTree as ET

    try:
        # 用 u2 直接取得 hierarchy（比 adb pull 可靠）
        xml_str = d.dump_hierarchy(compressed=False)
        # 同時存到檔案方便除錯
        xml_local = os.path.join(SCREENSHOT_DIR, "product_dump.xml")
        os.makedirs(SCREENSHOT_DIR, exist_ok=True)
        with open(xml_local, 'w', encoding='utf-8') as f:
            f.write(xml_str)

        root = ET.fromstring(xml_str)

        def _bounds_center(bounds_str):
            m = re.findall(r'\d+', bounds_str)
            if len(m) == 4:
                x1, y1, x2, y2 = map(int, m)
                return (x1 + x2) // 2, (y1 + y2) // 2
            return None

        def _find_clickable_containing(node, target_text):
            """找最小的 clickable 容器，其後代含有 target_text"""
            def _has_text(n, txt):
                if txt in n.attrib.get('text', '') or txt in n.attrib.get('content-desc', ''):
                    return True
                return any(_has_text(c, txt) for c in n)

            best = None  # (area, cx, cy, bounds_str)
            for n in root.iter('node'):
                if n.attrib.get('clickable') == 'true' and _has_text(n, target_text):
                    b = n.attrib.get('bounds', '')
                    m2 = re.findall(r'\d+', b)
                    if len(m2) == 4:
                        x1, y1, x2, y2 = map(int, m2)
                        area = (x2 - x1) * (y2 - y1)
                        if best is None or area < best[0]:
                            best = (area, (x1+x2)//2, (y1+y2)//2, b)
            if best:
                print(f"  [UI dump] 找到 clickable 容器(text='{target_text}') bounds={best[3]} center=({best[1]},{best[2]})")
                return (best[1], best[2])
            return None

        # 嘗試幾種文字找 clickable 容器（順序：最精確的先）
        result = None
        for search_text in ['點擊以新增商品', '新增商品']:
            result = _find_clickable_containing(root, search_text)
            if result:
                break
        if result is None:
            # 印出所有 clickable 元素輔助除錯
            print("  [UI dump] 找不到「新增商品」，所有 clickable 元素：")
            for node in root.iter('node'):
                if node.attrib.get('clickable') == 'true':
                    t = node.attrib.get('text', '')
                    d2 = node.attrib.get('content-desc', '')
                    b  = node.attrib.get('bounds', '')
                    if t or d2:
                        print(f"    text={t[:25]:25s} desc={d2[:25]:25s} bounds={b}")
        return result
    except Exception as e:
        print(f"  [UI dump] error: {e}")
        return None


def add_product(d, product_name):
    """搜尋並加入商品"""
    print(f"  新增商品: {product_name[:30]}...")

    # 先截圖確認在發布頁
    screenshot("publish_page_before_product")

    # ── 找「新增商品」可點擊區域 ──
    found_product_btn = False

    # 先嘗試直接 UI dump 找 clickable 父容器（最精確）
    coords = _dump_ui_find_product_btn(d)
    if coords:
        cx, cy = coords
        print(f"  → UI dump tap=({cx},{cy})")
        adb(f"shell input tap {cx} {cy}")
        time.sleep(5)
        screenshot("after_product_tap")
        found_product_btn = True
    else:
        # fallback: 嘗試滑動頁面後用 uiautomator2 找
        for attempt in range(4):
            btn = d(textContains='新增商品')
            if btn.exists(timeout=2):
                info = btn.info
                bounds = info.get('bounds', {})
                # 用元素自身 X 中心（不強制用 540，避免打到別的元素）
                cx = (bounds.get('left', 0) + bounds.get('right', 540)) // 2
                cy = (bounds.get('top', 0) + bounds.get('bottom', 100)) // 2
                print(f"  → u2 fallback bounds={bounds} tap=({cx},{cy})")
                adb(f"shell input tap {cx} {cy}")
                time.sleep(5)
                screenshot("after_product_tap")
                found_product_btn = True
                break
            try:
                scrollable = d(scrollable=True)
                if scrollable.exists(timeout=1):
                    scrollable.scroll.forward(steps=5)
            except Exception:
                swipe(540, 1500, 540, 800, 300, wait=1)
            time.sleep(1)

    if not found_product_btn:
        if "add_product" in COORD:
            print("  → 座標 fallback: add_product")
            adb(f"shell input tap {COORD['add_product'][0]} {COORD['add_product'][1]}")
            time.sleep(5)
            screenshot("add_product_coord")
        else:
            print("  ⚠ 找不到新增商品按鈕")
            return False

    # 確認是否進入商品搜尋頁（如果還在發布頁，表示 tap 沒作用）
    time.sleep(1)
    if (d(textContains='撰寫內文').exists(timeout=2)
            or d(text='發佈').exists(timeout=2)):
        # 還在發布頁，表示 tap 失敗 — 試一次 add_product 座標
        print("  ⚠ 仍在發布頁，tap 未生效，改用 add_product 座標")
        if "add_product" in COORD:
            adb(f"shell input tap {COORD['add_product'][0]} {COORD['add_product'][1]}")
            time.sleep(5)
            screenshot("after_product_tap2")
        else:
            print("  ⚠ 沒有 add_product 備用座標，跳過")
            return False

    screenshot("add_product_page")

    # ── 切換到「全部分類」tab（預設在「找讚好物」，找不到分潤商品）──
    switched_tab = False
    for _tab_text in ['全部分類', '全部', '所有商品']:
        _tab = d(textContains=_tab_text)
        if _tab.exists(timeout=2):
            _tab.click()
            print(f"  → 切換 tab：{_tab_text}")
            time.sleep(1.5)
            switched_tab = True
            break
    if not switched_tab:
        # React Native fallback：用座標點 tab
        _tx, _ty = COORD.get("all_categories_tab", (540, 230))
        print(f"  → ADB tap 全部分類 tab ({_tx},{_ty})")
        adb(f"shell input tap {_tx} {_ty}")
        time.sleep(1.5)
    screenshot("all_categories_tab")

    # 商品頁面完全 React Native，uiautomator2 看不到元素
    # 商品頁搜尋框座標（React Native，只能用 ADB 座標）
    _sx, _sy = COORD.get("product_search_box", (700, 130))
    print(f"  → 點搜尋框 ({_sx},{_sy})")
    adb(f"shell input tap {_sx} {_sy}")
    time.sleep(1.5)
    screenshot("search_box_focused")

    # 搜尋商品 - 嘗試不同長度的關鍵字
    search_terms = _build_search_terms(product_name)

    for term in search_terms:
        # 搜尋詞只剩 1 個詞時太模糊，跳過
        if len(term.split()) <= 1:
            print(f"    ⚠ 搜尋詞過短（{term}），跳過以免加錯商品")
            break

        print(f"    搜尋: {term}")
        # 每次先點搜尋框確保 focus（可能開 overlay）
        adb(f"shell input tap {_sx} {_sy}")
        time.sleep(1.5)

        # 找輸入框（overlay 可能是 native EditText）
        _input = None
        for _cls in ['android.widget.EditText', 'android.widget.AutoCompleteTextView']:
            _el = d(className=_cls)
            if _el.exists(timeout=2):
                _input = _el
                break
        if _input is None:
            _el = d(focused=True)
            if _el.exists(timeout=1):
                _input = _el

        if _input is not None:
            try:
                _input.clear_text()
                time.sleep(0.3)
                _input.set_text(term)   # 走 Accessibility，不用 clipboard
                time.sleep(0.8)
            except Exception as _se:
                print(f"    ⚠ set_text 失敗: {_se}，跳過此詞")
                continue
        else:
            print(f"    ⚠ 找不到輸入框，跳過此詞")
            continue

        adb("shell input keyevent KEYCODE_ENTER")
        time.sleep(3)
        screenshot(f"search_result")

        # 先試 uiautomator2（有時 React Native 會暴露 text）
        no_result = d(textContains='沒有搜尋結果')
        if no_result.exists(timeout=2):
            print(f"    ✗ 沒結果，縮短再試")
            # 重新點搜尋框準備下一輪
            adb(f"shell input tap {_sx} {_sy}")
            time.sleep(0.5)
            continue

        add_btn = d(text='加入')
        if add_btn.exists(timeout=3):
            add_btn[0].click()
            print("    ✓ 已點加入 (u2)")
            time.sleep(3)
        else:
            # React Native fallback：用 ADB tap 點第一個「加入」按鈕座標
            _jx, _jy = COORD.get("product_join_btn", (960, 340))
            print(f"    → ADB tap 加入 ({_jx},{_jy})")
            adb(f"shell input tap {_jx} {_jy}")
            time.sleep(3)
            screenshot("after_join_tap")

            # 確認是否成功進入商品編輯頁（顯示「完成」或「確認」）
            if not (d(text='完成').exists(timeout=3)
                    or d(text='確認').exists(timeout=2)):
                # 還在搜尋頁，可能沒搜到商品
                print(f"    ✗ 無法確認「加入」成功，縮短搜尋詞再試")
                # 重新點搜尋框
                adb(f"shell input tap {_sx} {_sy}")
                time.sleep(0.5)
                continue

        # ── 到這裡表示成功加入商品 ──
        # 編輯商品資訊頁 → 點完成
        done = d(text='完成')
        if done.exists(timeout=5):
            done.click()
            time.sleep(2)
            screenshot("after_product_done")

        # 返回發布頁
        d.press('back')
        time.sleep(2)
        if d(textContains='新增商品').exists(timeout=1):
            d.press('back')
            time.sleep(2)

        print("  ✓ 商品已加入")
        return True

    print("  ⚠ 找不到商品，返回發布頁")
    _navigate_back_to_publish(d)
    return False


def _build_search_terms(product_name):
    """根據品名建立由短到長的搜尋詞（品牌+品名優先）"""
    # 去掉 emoji 和特殊符號
    clean = re.sub(r'[^\w\s]', ' ', product_name)
    clean = re.sub(r'\s+', ' ', clean).strip()
    words = clean.split()

    # 跳過太通用的詞
    skip_words = {'現貨', '免運', '台灣', '出貨', '限時', '特價', '熱賣', '新款',
                   '隔日達', '當日', '預購', '批發', '包郵', '直送', '即日',
                   '近日', '到貨', '工廠', '直營', '正品', '爆款', '全年無休',
                   '附發票', '免運費'}
    filtered = [w for w in words if w not in skip_words]

    terms = []
    # 優先用品牌+品名（前2個有意義的詞）
    if len(filtered) >= 2:
        terms.append(' '.join(filtered[:2]))
    # 再試前3個詞
    if len(filtered) >= 3:
        terms.append(' '.join(filtered[:3]))
    # 再試前4個詞
    if len(filtered) >= 4:
        terms.append(' '.join(filtered[:4]))
    # 最後只用第一個詞（品牌名）
    if len(filtered) >= 1:
        terms.append(filtered[0])

    # 如果全部被過濾掉，用原始的
    if not terms and words:
        terms = [' '.join(words[:2]), words[0]]

    return terms


def _navigate_back_to_publish(d):
    """安全返回到發布頁"""
    for _ in range(5):
        # 已經在發布頁？
        if d(textContains='撰寫內文').exists(timeout=1):
            return
        if d(textContains='發佈').exists(timeout=1):
            return
        # 有捨棄對話框？點取消留在頁面
        cancel = d(text='取消')
        if cancel.exists(timeout=1):
            cancel.click()
            time.sleep(1)
            return
        # 還在新增商品頁？按返回
        d.press('back')
        time.sleep(2)


def publish(d, dry_run=False):
    """點發佈"""
    screenshot("before_publish")

    if dry_run:
        print("  🔸 DRY-RUN: 到此為止，不真的發佈")
        # 按返回離開
        d.press('back')
        time.sleep(1)
        discard = d(text='Discard')
        if not discard.exists(timeout=2):
            discard = d(textContains='捨棄')
        if discard.exists(timeout=2):
            discard.click()
            time.sleep(2)
        return True

    print("  發佈中...")
    # 嘗試多個按鈕文字（繁體）
    for txt in ['發佈', '發布', 'Publish']:
        btn = d(text=txt)
        if btn.exists(timeout=3):
            btn.click()
            time.sleep(5)
            return True

    print("  ⚠ 找不到發佈按鈕")
    screenshot("publish_fail")
    return False


def handle_upload_failure(d, max_retries=3):
    """處理上傳失敗：關閉全部分頁 → 重開蝦皮 → 自動重傳"""
    for attempt in range(max_retries):
        print(f"  檢查上傳狀態（嘗試 {attempt+1}/{max_retries}）...")
        time.sleep(5)
        screenshot(f"upload_check_{attempt}")

        # 檢查是否上傳失敗
        fail = d(textContains='上傳失敗')
        if fail.exists(timeout=3):
            print("  ⚠ 上傳失敗，關閉全部分頁重試...")
            close_all_apps()
            open_shopee()
            time.sleep(3)
            # 進短影音 feed 觸發自動重傳
            tap(*COORD["shortvideo_tab"], wait=5)

            # 等待上傳完成
            for i in range(30):  # 最多等 60 秒
                time.sleep(2)
                uploading = d(textContains='上傳中')
                if uploading.exists:
                    print(f"    上傳中...")
                    continue
                fail2 = d(textContains='上傳失敗')
                if fail2.exists:
                    print("    又失敗了，再試...")
                    break
                # 沒有上傳中也沒有失敗 = 成功
                print("  ✓ 上傳完成！")
                return True

        else:
            # 檢查是否正在上傳中
            uploading = d(textContains='上傳中')
            if uploading.exists:
                print("  上傳中，等待完成...")
                for i in range(30):
                    time.sleep(2)
                    if not d(textContains='上傳中').exists(timeout=2):
                        if not d(textContains='上傳失敗').exists(timeout=2):
                            print("  ✓ 上傳完成！")
                            return True
                        break
            else:
                print("  ✓ 上傳完成！")
                return True

    print("  ✗ 多次重試仍失敗")
    return False


def upload_one(row_data, dry_run=False):
    """上傳一部影片的完整流程"""
    idx = row_data['編號']
    excel_row = row_data['excel_row']  # Excel 行號，對應檔名
    caption = row_data.get('標題', '') or row_data.get('文案', '') or row_data.get('關鍵字文案', '') or ''
    _matches = glob.glob(os.path.join(VIDEO_DIR, f"{excel_row-1:03d}_*.mp4"))
    video_path = _matches[0] if _matches else os.path.join(VIDEO_DIR, f"{excel_row-1:03d}_final.mp4")
    # 用影片檔名的品名做商品搜尋（避免 Excel 新增列後 row 號碼與品名錯位）
    _fname = os.path.basename(video_path)
    _m = re.match(r'^\d+_(.+?)\.mp4$', _fname, re.IGNORECASE)
    name = _m.group(1).strip() if _m else row_data['品名']

    print(f"\n{'='*50}")
    print(f"📹 上傳第 {idx} 部: {name[:30]}...")
    print(f"{'='*50}")

    # 每支影片開始前自動解鎖螢幕
    unlock_screen(pin="0000")

    # 檢查影片是否存在
    if not os.path.exists(video_path):
        print(f"  ✗ 影片不存在: {video_path}")
        return False

    if not caption:
        print(f"  ✗ 沒有文案，跳過")
        return False

    # 1. 推影片到手機
    remote_path = push_video(video_path)

    # 2. 連接 u2
    global _d
    d = u2.connect(DEVICE)
    _d = d  # 讓 tap() 可以用 u2 HTTP API

    # 3. 開啟蝦皮
    open_shopee()

    # 4. 導航到拍攝頁
    navigate_to_create(d)

    # 5. 從媒體庫選影片
    select_video_from_gallery(d)

    # 6. 跳過編輯頁
    skip_editor(d)

    # 7. 輸入文案
    enter_caption(d, caption)

    # 8. 關閉合拍/拼接
    toggle_off_switches(d)

    # 9. 新增商品
    add_product(d, name)

    # 確保回到發布頁
    _navigate_back_to_publish(d)

    # 10. 發佈
    result = publish(d, dry_run=dry_run)

    if not dry_run and result:
        # 11. 處理上傳結果
        success = handle_upload_failure(d)
        # 12. 清掉手機上的影片
        if success:
            cleanup_phone_video(remote_path)
        return success

    # dry-run 也清掉影片
    cleanup_phone_video(remote_path)
    return result


def read_excel():
    """讀取 Excel 資料，excel_row = 實際 Excel 行號（從2開始），用於對應影片檔名"""
    wb = openpyxl.load_workbook(EXCEL_PATH)
    ws = wb.active

    _index_json = os.getenv("EXCEL_INDEX", "").strip()
    if _index_json:
        # EXCEL_INDEX 格式：JSON {"編號":0,"品名":2,"關鍵字文案":5}（0-based 欄位索引）
        _col_map = json.loads(_index_json)
        rows = []
        for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
            data = {k: (row[v] if v < len(row) else None) for k, v in _col_map.items()}
            if data.get('編號') and data.get('品名'):
                data['excel_row'] = row_idx
                rows.append(data)
        return rows

    # 預設：用第一列標題自動對應（家裡電腦 / 標準格式）
    headers = [cell.value for cell in ws[1]]
    rows = []
    for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        data = dict(zip(headers, row))
        if data.get('編號') and data.get('品名'):
            data['excel_row'] = row_idx  # Excel 行號，對應 row002, row003...
            rows.append(data)
    return rows


def main():
    global DEVICE, COORD

    parser = argparse.ArgumentParser(description='蝦皮短影音批次上傳')
    parser.add_argument('--row',    type=int,  help='只上傳指定 Excel 行號（如 --row 3 = row003_final.mp4）')
    parser.add_argument('--dry-run', action='store_true', help='跑到發佈前停下')
    parser.add_argument('--start',  type=int,  default=1,          help='從第幾筆開始')
    parser.add_argument('--count',  type=int,  default=50,         help='上傳幾部')
    parser.add_argument('--device', type=str,  default=None,        help='手機 IP:PORT（不填則從 profile 自動取得）')
    parser.add_argument('--phone',  type=str,  default="a52s",
                        choices=list(PHONE_PROFILES.keys()),
                        help='手機型號 profile：mi_note2 / s25fe / a52s / custom（預設: a52s）')
    args = parser.parse_args()

    COORD  = PHONE_PROFILES[args.phone]
    # 裝置 IP：--device 優先；沒有則從 profile 取；最後用模組預設值
    if args.device:
        DEVICE = args.device
    elif "device" in COORD:
        DEVICE = COORD["device"]
    # 否則維持模組層的 DEVICE 預設值

    print("🚀 蝦皮短影音批次上傳")
    print(f"   裝置: {DEVICE}")
    print(f"   手機 Profile: {args.phone}")
    print(f"   Dry-run: {args.dry_run}")

    # 連接裝置
    connect_device()

    # 讀取 Excel
    rows = read_excel()
    print(f"   Excel 共 {len(rows)} 筆資料")

    # 篩選要上傳的
    if args.row:
        targets = [r for r in rows if r['excel_row'] == args.row]
    else:
        # 只選有影片檔案的
        targets = []
        for r in rows:
            if r['excel_row'] < args.start:
                continue
            _m = glob.glob(os.path.join(VIDEO_DIR, f"{r['excel_row']-1:03d}_*.mp4"))
            video_path = _m[0] if _m else ""
            if video_path and (r.get('標題') or r.get('文案') or r.get('關鍵字文案')):
                targets.append(r)
            if len(targets) >= args.count:
                break

    print(f"   本次上傳: {len(targets)} 部")
    print()

    success_count = 0
    fail_count = 0

    for i, row_data in enumerate(targets):
        print(f"\n[{i+1}/{len(targets)}]", end="")
        try:
            ok = upload_one(row_data, dry_run=args.dry_run)
            if ok:
                success_count += 1
                print(f"  ✅ 第 {row_data['編號']} 部完成")
            else:
                fail_count += 1
                print(f"  ❌ 第 {row_data['編號']} 部失敗")
        except Exception as e:
            fail_count += 1
            print(f"  ❌ 第 {row_data['編號']} 部異常: {e}")
            # 嘗試回到安全狀態
            try:
                close_all_apps()
            except:
                pass

    print(f"\n{'='*50}")
    print(f"📊 結果: 成功 {success_count} / 失敗 {fail_count} / 總共 {len(targets)}")
    print(f"{'='*50}")


if __name__ == '__main__':
    main()
