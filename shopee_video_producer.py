# 蝦皮影片後製腳本 v2
# 功能：_clips_XXX → resize → 合併 → 旁白TTS逐句 + 字幕同步 + BGM → output_final
# 影片內不加任何標題文字，標題只用於上傳平台

import sys, io, os, re, glob, time, shutil, tempfile, subprocess, random, logging, asyncio
sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8', errors='replace')
sys.stderr = io.TextIOWrapper(sys.stderr.buffer, encoding='utf-8', errors='replace')

from dotenv import load_dotenv
load_dotenv(dotenv_path=os.path.join(os.path.dirname(os.path.abspath(__file__)), '.env'))

import numpy as np
import openpyxl
from PIL import Image, ImageDraw, ImageFont
from moviepy import (VideoFileClip, AudioFileClip, ImageClip,
                     concatenate_videoclips, concatenate_audioclips,
                     CompositeVideoClip, CompositeAudioClip)
import imageio_ffmpeg
import edge_tts
import gc

logging.basicConfig(level=logging.INFO,
    format='%(asctime)s [%(levelname)s] %(message)s', datefmt='%H:%M:%S')
log = logging.getLogger(__name__)

# ── 路徑設定（從 .env 讀）────────────────────────────────────────────────────
EXCEL_PATH  = os.getenv('EXCEL_PATH',  r'D:\Users\user\Desktop\蝦皮影片專案\蝦皮關鍵字選品_2026年3-4月new.xlsx')
CLIPS_DIR   = os.getenv('OUTPUT_DIR',  r'D:\Users\user\Desktop\蝦皮影片專案\output_videos')
FINAL_DIR   = os.getenv('FINAL_DIR',   r'D:\Users\user\Desktop\蝦皮影片專案\output_final')
BGM_DIR     = os.getenv('BGM_DIR',     r'D:\Users\user\Desktop\蝦皮影片專案\music')
FONT_PATH   = os.getenv('FONT_PATH',   r'C:\Windows\Fonts\msjh.ttc')
FFMPEG_EXE  = os.getenv('FFMPEG_PATH', '') or imageio_ffmpeg.get_ffmpeg_exe()

# ── Excel 欄位 ────────────────────────────────────────────────────────────────
COL_NAME   = int(os.getenv('COL_NAME',   '2'))
COL_COPY   = int(os.getenv('COL_COPY',   '8'))
COL_TITLE  = int(os.getenv('COL_TITLE',  '9'))
COL_STATUS = int(os.getenv('COL_STATUS', '10'))

# ── 影片參數 ──────────────────────────────────────────────────────────────────
MIN_CLIPS        = 1
VIDEO_W, VIDEO_H = 1080, 1920
TTS_VOICE        = os.getenv('TTS_VOICE', 'zh-TW-HsiaoChenNeural')  # 女聲；換男聲用 zh-TW-YunJheNeural

# ══════════════════════════════════════════════════════════════════════════════
# 斷句
# ══════════════════════════════════════════════════════════════════════════════
def split_sentences(text):
    """將文案拆成逐句 list，以中文句號/驚嘆號/問號為斷點"""
    text = str(text).replace('\\n', '\n').replace('\n', '。')
    parts = re.split(r'(?<=[。！？])', text)
    # 過濾：必須含至少一個中文字，避免空句/純標點/emoji 讓 TTS 失敗
    return [s.strip() for s in parts
            if s.strip() and re.search(r'[\u4e00-\u9fff]', s)]

# ══════════════════════════════════════════════════════════════════════════════
# TTS 生成
# ══════════════════════════════════════════════════════════════════════════════
async def _tts_one(i, text, path):
    """單句 TTS，失敗最多重試 3 次"""
    for attempt in range(3):
        try:
            await edge_tts.Communicate(text, TTS_VOICE).save(path)
            return i, path, True
        except Exception as e:
            log.warning('  TTS [%d] 第%d次失敗: %s', i, attempt + 1, e)
            await asyncio.sleep(1.5 * (attempt + 1))
    return i, path, False

async def _build_tts_async(sentences, tmpdir):
    """所有句子同時送出（網路 I/O，不吃 CPU）"""
    tasks = [
        _tts_one(i, sent, os.path.join(tmpdir, f'tts_{i:03d}.mp3'))
        for i, sent in enumerate(sentences)
    ]
    return await asyncio.gather(*tasks)

def build_tts_segments(sentences, tmpdir):
    """
    並行生成所有句 TTS MP3，測量時長
    回傳 list of dict: {text, audio_path, duration}
    """
    results = asyncio.run(_build_tts_async(sentences, tmpdir))
    # results 順序與 sentences 一致（gather 保序）
    segments = []
    for i, (_, path, ok) in enumerate(sorted(results, key=lambda x: x[0])):
        sent = sentences[i]
        if ok and os.path.exists(path):
            clip = AudioFileClip(path)
            dur = clip.duration
            clip.close()
            segments.append({'text': sent, 'audio': path, 'duration': dur})
            log.info('  TTS [%d] %.1fs → %s', i, dur, sent[:30])
        else:
            segments.append({'text': sent, 'audio': None, 'duration': 2.5})
    return segments

# ══════════════════════════════════════════════════════════════════════════════
# 字幕圖層（帶 start_time）
# ══════════════════════════════════════════════════════════════════════════════
def make_subtitle_clip(text, start_time, duration, font_size=52, bg_alpha=170):
    try:
        font = ImageFont.truetype(FONT_PATH, font_size)
    except Exception:
        font = ImageFont.load_default()

    margin = 40
    dummy = ImageDraw.Draw(Image.new('RGBA', (1, 1)))
    lines = []
    for paragraph in str(text).split('\n'):
        current = ''
        for ch in paragraph:
            test = current + ch
            if dummy.textlength(test, font=font) > VIDEO_W - margin * 2 and current:
                lines.append(current)
                current = ch
            else:
                current = test
        if current:
            lines.append(current)

    if not lines:
        return None

    line_h    = font_size + 14
    box_h     = line_h * len(lines) + margin
    h_pad     = 24
    max_text_w = int(max(dummy.textlength(line, font=font) for line in lines))
    box_w     = min(max_text_w + h_pad * 2, VIDEO_W)
    box_x     = (VIDEO_W - box_w) // 2
    img       = Image.new('RGBA', (VIDEO_W, box_h), (0, 0, 0, 0))
    draw      = ImageDraw.Draw(img)
    draw.rectangle([box_x, 0, box_x + box_w, box_h], fill=(0, 0, 0, bg_alpha))

    for i, line in enumerate(lines):
        w = draw.textlength(line, font=font)
        x = (VIDEO_W - w) // 2
        y = margin // 2 + i * line_h
        draw.text((x + 2, y + 2), line, font=font, fill=(0, 0, 0, 200))
        draw.text((x, y), line, font=font, fill=(255, 255, 255, 255))

    arr   = np.array(img)
    pos_y = (VIDEO_H - box_h) // 2  # 垂直置中

    return (ImageClip(arr, transparent=True)
            .with_duration(duration)
            .with_start(start_time)
            .with_position(('center', pos_y)))

# ══════════════════════════════════════════════════════════════════════════════
# BGM
# ══════════════════════════════════════════════════════════════════════════════
def get_bgm(duration):
    files = []
    for ext in ('*.mp3', '*.wav', '*.m4a', '*.aac'):
        files += glob.glob(os.path.join(BGM_DIR, ext))
    if not files:
        return None
    try:
        audio = AudioFileClip(random.choice(files))
        if audio.duration < duration:
            loops = int(duration / audio.duration) + 1
            audio = concatenate_audioclips([audio] * loops)
        return audio.subclipped(0, duration).with_volume_scaled(0.25)
    except Exception as e:
        log.warning('[BGM] %s', e)
        return None

# ══════════════════════════════════════════════════════════════════════════════
# 後製主函式
# ══════════════════════════════════════════════════════════════════════════════
def produce_video(clip_paths, copy_text, output_path):
    with tempfile.TemporaryDirectory(ignore_cleanup_errors=True) as tmpdir:

        # ── Step 1: FFmpeg resize 到 1080x1920 ──────────────────────────────
        resized = []
        for i, src in enumerate(clip_paths):
            out = os.path.join(tmpdir, f'clip{i:02d}.mp4')
            cmd = [FFMPEG_EXE, '-y', '-i', src,
                   '-vf', (f'scale={VIDEO_W}:{VIDEO_H}:force_original_aspect_ratio=decrease,'
                            f'pad={VIDEO_W}:{VIDEO_H}:(ow-iw)/2:(oh-ih)/2:black'),
                   '-c:v', 'libx264', '-an', '-preset', 'fast', '-loglevel', 'error', out]
            subprocess.run(cmd, capture_output=True, timeout=120)
            if os.path.exists(out):
                resized.append(out)

        if len(resized) < MIN_CLIPS:
            log.warning('resize 後只有 %d 支，跳過', len(resized))
            return False

        # ── Step 2: 斷句 + 生成 TTS ─────────────────────────────────────────
        has_copy = copy_text and str(copy_text).strip() not in ('', 'None', '文案')
        segments = []
        total_tts_dur = 0.0

        if has_copy:
            sentences = split_sentences(str(copy_text))
            log.info('  斷句 %d 句', len(sentences))
            segments = build_tts_segments(sentences, tmpdir)
            total_tts_dur = sum(s['duration'] for s in segments)
            log.info('  TTS 總時長 %.1fs', total_tts_dur)

        # ── Step 3: 載入影片，循環補足 TTS 時長 ─────────────────────────────
        try:
            raw_clips = [VideoFileClip(p).without_audio() for p in resized]
            merged    = concatenate_videoclips(raw_clips, method='compose')
            raw_dur   = merged.duration

            target_dur = max(total_tts_dur + 1.5, raw_dur)  # TTS 結束後留 1.5s

            if raw_dur < target_dur:
                loops  = int(target_dur / raw_dur) + 1
                merged = concatenate_videoclips([merged] * loops, method='compose')

            merged = merged.subclipped(0, target_dur)
            dur    = merged.duration

        except Exception as e:
            log.error('[video load] %s', e)
            return False

        # ── Step 4: 字幕圖層（逐句，帶 start_time）──────────────────────────
        layers = [merged]
        t = 0.0
        for seg in segments:
            ov = make_subtitle_clip(seg['text'], t, seg['duration'])
            if ov:
                layers.append(ov)
            t += seg['duration']

        # ── Step 5: 組合音訊（TTS + BGM）────────────────────────────────────
        try:
            tts_clips = [AudioFileClip(s['audio'])
                         for s in segments if s['audio'] and os.path.exists(s['audio'])]

            bgm = get_bgm(dur)

            if tts_clips:
                tts_audio = concatenate_audioclips(tts_clips)
                if bgm:
                    mixed_audio = CompositeAudioClip([
                        tts_audio.with_volume_scaled(1.0),
                        bgm
                    ])
                    log.info('  TTS + BGM 合成')
                else:
                    mixed_audio = tts_audio
                    log.info('  只有 TTS（無 BGM）')
            elif bgm:
                mixed_audio = bgm
                log.info('  只有 BGM（無 TTS）')
            else:
                mixed_audio = None

        except Exception as e:
            log.warning('[audio] %s', e)
            mixed_audio = None

        # ── Step 6: 合成輸出 ─────────────────────────────────────────────────
        try:
            final = CompositeVideoClip(layers, size=(VIDEO_W, VIDEO_H))
            if mixed_audio:
                final = final.with_audio(mixed_audio)

            final.write_videofile(output_path, codec='libx264', audio_codec='aac',
                                  fps=30, logger=None,
                                  ffmpeg_params=['-loglevel', 'error', '-preset', 'veryfast'])

            for c in raw_clips: c.close()
            merged.close()
            if tts_clips:
                for c in tts_clips: c.close()
            if bgm: bgm.close()
            final.close()
            gc.collect()
            return True

        except Exception as e:
            log.error('[compose] %s', e)
            return False

# ══════════════════════════════════════════════════════════════════════════════
# 主程式
# ══════════════════════════════════════════════════════════════════════════════
def main():
    import argparse
    parser = argparse.ArgumentParser()
    parser.add_argument('--start', type=int, default=None, help='從第幾列開始（Excel row，含標題行，預設從第2列）')
    parser.add_argument('--count', type=int, default=None, help='最多後製幾部（預設不限）')
    args = parser.parse_args()

    os.makedirs(FINAL_DIR, exist_ok=True)

    wb    = openpyxl.load_workbook(EXCEL_PATH)
    ws    = wb.active
    total = ws.max_row - 1
    done, fail, skip = 0, 0, 0

    start_row = args.start if args.start else 2
    max_count = args.count  # None = 不限

    print('=' * 55)
    print('蝦皮影片後製 v2（旁白TTS + 逐句字幕）')
    print(f'clips 來源: {CLIPS_DIR}')
    print(f'輸出:       {FINAL_DIR}')
    print(f'TTS 聲音:   {TTS_VOICE}')
    print(f'起始列: {start_row}，最多後製: {max_count if max_count else "不限"} 部')
    print('=' * 55)

    for row in ws.iter_rows(min_row=start_row):
        row_idx     = row[0].row - 1
        name        = row[COL_NAME   - 1].value
        copy_text   = row[COL_COPY   - 1].value
        status_cell = row[COL_STATUS - 1]

        if not name:
            continue

        if status_cell.value == '影片完成':
            skip += 1
            continue

        # 跳過黃色列（品名欄 FFFFFF00 = 直播/垃圾）
        name_cell = row[COL_NAME - 1]
        name_fill = name_cell.fill
        if name_fill and name_fill.fgColor and name_fill.fgColor.type == 'rgb':
            if name_fill.fgColor.rgb == 'FFFFFF00':
                skip += 1
                continue

        # 跳過沒有文案的列（避免產出無旁白影片）
        if not copy_text:
            continue

        clips_folder = os.path.join(CLIPS_DIR, f'_clips_{row_idx:03d}')
        if not os.path.exists(clips_folder):
            continue

        clip_files = sorted([
            os.path.join(clips_folder, f)
            for f in os.listdir(clips_folder) if f.endswith('.mp4')
        ])

        if len(clip_files) < MIN_CLIPS:
            log.info('[skip %d] clip 不足 %d 支', row_idx, len(clip_files))
            continue

        safe     = re.sub(r'[\\/:*?"<>|]', '', str(name))[:35]
        out_path = os.path.join(FINAL_DIR, f'{row_idx:03d}_{safe}.mp4')

        if os.path.exists(out_path):
            log.info('[skip %d] 已存在', row_idx)
            status_cell.value = '影片完成'
            skip += 1
            continue

        print(f'\n[{row_idx:3}/{total}] {str(name)[:45]}')
        if produce_video(clip_files, copy_text, out_path):
            status_cell.value = '影片完成'
            done += 1
            log.info('完成 → %s', os.path.basename(out_path))
        else:
            status_cell.value = '後製失敗'
            fail += 1
            log.warning('失敗 #%d', row_idx)

        wb.save(EXCEL_PATH)

        # --count 達到上限就停
        if max_count and done >= max_count:
            print(f'\n已完成指定數量 {max_count} 部，停止。')
            break

    print(f'\n{"=" * 55}')
    print(f'完成 {done} 個 / 失敗 {fail} 個 / 跳過 {skip} 個')
    print(f'輸出: {FINAL_DIR}')

if __name__ == '__main__':
    main()
