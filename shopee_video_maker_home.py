# [HOME COMPUTER VERSION] 蝦皮評論影片下載 + Gemini過濾 + 合併
# 從選品Excel → 抓每個商品評論區影片 → Gemini篩掉人臉/開箱 → 湊3支合一
# 建立在 video_filter.py (Gemini) + shopee_keyword_scraper.py (CDP) 基礎上

import sys, io, asyncio, json, re, os, random, requests, time, tempfile, shutil
sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8', errors='replace')
sys.stderr = io.TextIOWrapper(sys.stderr.buffer, encoding='utf-8', errors='replace')

import cv2
import warnings
warnings.filterwarnings('ignore')
import google.generativeai as genai
import openpyxl
from playwright.async_api import async_playwright
from moviepy import VideoFileClip, concatenate_videoclips
import imageio_ffmpeg

# ── Config ────────────────────────────────────────────────────────────────────
EXCEL_PATH   = r'C:\Users\user\Desktop\蝦皮關鍵字選品_2026年3-4月.xlsx'
OUTPUT_DIR   = r'C:\Users\user\Desktop\蝦皮選品影片'
GEMINI_KEY   = 'AIzaSyBfhyW5K3TrNZHs5380tBQ2KjabCmXHtW'  # same as video_filter.py

MIN_CLIPS    = 3    # 每個商品最少要幾支影片才合併
MAX_TRY      = 15   # 每個商品最多試幾支評論影片
RATINGS_LIMIT = 10  # 每頁抓幾筆評論
TARGET_SEC   = 18   # 合併後目標秒數

COL_NAME   = 1   # A: 品名
COL_LINK   = 2   # B: 分潤連結 (含 shop_id / item_id)
COL_STATUS = 8   # H: 狀態

FFMPEG_EXE = imageio_ffmpeg.get_ffmpeg_exe()

# ── Gemini 初始化 (沿用 video_filter.py 設定) ─────────────────────────────────
genai.configure(api_key=GEMINI_KEY, client_options={'api_version': 'v1'})
gemini = genai.GenerativeModel(model_name='gemini-1.5-flash')

# ── 工具函式 ──────────────────────────────────────────────────────────────────
def extract_ids(url):
    """從商品連結抓出 shop_id, item_id"""
    s = str(url or '')
    m = re.search(r'/product/(\d+)/(\d+)', s)
    if m: return m.group(1), m.group(2)
    m = re.search(r'-i\.(\d+)\.(\d+)', s)
    if m: return m.group(1), m.group(2)
    return None, None

def gemini_check_frame(img_path):
    """
    用 Gemini 檢查一張截圖，一次問兩件事：
    1) 有無清楚人臉  2) 是否開箱影片（出現紙箱）
    回傳 (has_face: bool, is_unboxing: bool)
    """
    try:
        sample = genai.upload_file(path=img_path)
        resp = gemini.generate_content([
            sample,
            "請回答以下兩個問題，每個只能回答「是」或「否」，\n"
            "格式範例：否,否\n"
            "問題1：這張圖片中是否有清楚可辨識的人臉？\n"
            "問題2：這張圖片是否為開箱影片（有紙箱或包裝盒正在被拆開）？"
        ])
        text = resp.text.strip().replace('，', ',').replace(' ', '')
        parts = text.split(',')
        face     = len(parts) > 0 and '是' in parts[0]
        unboxing = len(parts) > 1 and '是' in parts[1]
        return face, unboxing
    except Exception as e:
        print(f'      [Gemini] {e}')
        return False, False   # 判斷失敗就放行（寧可留著）

def is_valid_video(video_path):
    """
    抽 3 個時間點的截圖 → 送 Gemini 檢查
    只要任何一張：有人臉 OR 是開箱 → 拒絕
    全部通過 → 接受
    """
    try:
        cap = cv2.VideoCapture(video_path)
        total = int(cap.get(cv2.CAP_PROP_FRAME_COUNT))
        if total == 0:
            cap.release(); return False

        for pct in (0.15, 0.50, 0.85):
            cap.set(cv2.CAP_PROP_POS_FRAMES, int(total * pct))
            ok, frame = cap.read()
            if not ok: continue

            tmp_img = video_path + f'_f{int(pct*100)}.jpg'
            cv2.imwrite(tmp_img, frame)
            face, unbox = gemini_check_frame(tmp_img)
            if os.path.exists(tmp_img): os.remove(tmp_img)

            if face:
                print(f'      X 人臉 @{int(pct*100)}%')
                cap.release(); return False
            if unbox:
                print(f'      X 開箱 @{int(pct*100)}%')
                cap.release(); return False

        cap.release()
        return True
    except Exception as e:
        print(f'      [valid_check] {e}')
        return False

def download_video(url, dest):
    try:
        r = requests.get(url, stream=True, timeout=30,
                         headers={'User-Agent': 'Mozilla/5.0',
                                  'Referer': 'https://shopee.tw/'})
        r.raise_for_status()
        with open(dest, 'wb') as f:
            for chunk in r.iter_content(65536): f.write(chunk)
        size = os.path.getsize(dest)
        return size > 5000   # 至少 5KB 才算有效
    except Exception as e:
        print(f'      [download] {e}')
        return False

def merge_clips(clip_paths, output_path):
    """
    用 moviepy v2 合併 3 支影片（去音訊），
    輸出到 output_path
    """
    try:
        clips = []
        for p in clip_paths[:MIN_CLIPS]:
            c = VideoFileClip(p).without_audio()
            clips.append(c)

        final = concatenate_videoclips(clips, method='compose')
        # 控制最長不超過 TARGET_SEC
        if final.duration > TARGET_SEC:
            final = final.subclipped(0, TARGET_SEC)

        final.write_videofile(
            output_path, codec='libx264', audio=False,
            logger=None,
            ffmpeg_params=['-loglevel', 'error', '-preset', 'fast']
        )
        for c in clips: c.close()
        final.close()
        return True
    except Exception as e:
        print(f'      [merge] {e}')
        return False

# ── 評論影片 URL 抓取（CDP fetch on shopee.tw） ───────────────────────────────
async def fetch_review_videos(page, shop_id, item_id):
    """
    呼叫 Shopee ratings API，回傳影片 URL 列表
    用 CDP page.evaluate() 內的 fetch，自動帶入登入 session
    """
    video_urls = []

    for offset in range(0, RATINGS_LIMIT * 30, RATINGS_LIMIT):
        js = f"""
        (async () => {{
          try {{
            const r = await fetch(
              '/api/v2/item/get_ratings?itemid={item_id}&shopid={shop_id}' +
              '&type=0&offset={offset}&limit={RATINGS_LIMIT}&filter=0',
              {{credentials: 'include'}}
            );
            return await r.json();
          }} catch(e) {{ return {{error: e.toString()}}; }}
        }})()
        """
        data = await page.evaluate(js)
        if not data or data.get('error'): break

        ratings = (data.get('data') or {}).get('ratings') or []
        if not ratings: break

        for rating in ratings:
            for v in (rating.get('videos') or []):
                url = v.get('url') or v.get('video_url') or v.get('download_url') or ''
                if url.startswith('http'): video_urls.append(url)
            for m in (rating.get('medias') or []):
                if m.get('type') == 2:
                    url = m.get('url') or m.get('video_url') or ''
                    if url.startswith('http'): video_urls.append(url)

        if len(ratings) < RATINGS_LIMIT: break   # 最後一頁
        if len(video_urls) >= MAX_TRY: break
        await asyncio.sleep(0.3)

    return list(dict.fromkeys(video_urls))[:MAX_TRY]  # deduplicate

# ── 處理單一商品 ─────────────────────────────────────────────────────────────
async def process_product(page, name, link, row_idx):
    shop_id, item_id = extract_ids(link)
    if not shop_id:
        print(f'  X 無法解析 ID: {str(link)[:60]}')
        return '無法解析ID'

    print(f'  shop={shop_id}  item={item_id}')

    # 如果目前頁面在驗證頁，先導到商品頁讓 session 生效
    current_url = page.url
    if 'captcha' in current_url or 'verify' in current_url:
        await page.goto(f'https://shopee.tw', wait_until='domcontentloaded', timeout=15000)
        await asyncio.sleep(2)

    video_urls = await fetch_review_videos(page, shop_id, item_id)
    print(f'  評論影片 URL: {len(video_urls)} 個')

    if not video_urls:
        return '無評論影片'

    tmpdir = tempfile.mkdtemp()
    valid_clips = []

    try:
        for i, url in enumerate(video_urls):
            if len(valid_clips) >= MIN_CLIPS: break

            dest = os.path.join(tmpdir, f'clip_{i:02d}.mp4')
            print(f'  [{i+1:2}/{len(video_urls)}] 下載...', end=' ', flush=True)

            if not download_video(url, dest):
                print('失敗')
                continue

            print('Gemini...', end=' ', flush=True)
            if is_valid_video(dest):
                valid_clips.append(dest)
                print(f'通過 ({len(valid_clips)}/{MIN_CLIPS})')
            # 拒絕原因已在 is_valid_video 內印出

            time.sleep(0.5)

        print(f'  有效: {len(valid_clips)} / 需要: {MIN_CLIPS}')

        if len(valid_clips) >= MIN_CLIPS:
            safe = re.sub(r'[\\/:*?"<>|]', '', str(name))[:35]
            out  = os.path.join(OUTPUT_DIR, f'{row_idx:03d}_{safe}.mp4')
            print(f'  合併 → {os.path.basename(out)} ...', end=' ', flush=True)
            if merge_clips(valid_clips, out):
                print('完成')
                return '影片完成'
            else:
                return '合併失敗'
        else:
            return f'影片不足({len(valid_clips)}/{MIN_CLIPS})'

    finally:
        shutil.rmtree(tmpdir, ignore_errors=True)

# ── 主程式 ────────────────────────────────────────────────────────────────────
async def main():
    print('=' * 55)
    print('蝦皮選品影片製作 - Home Computer Version')
    print(f'輸出資料夾: {OUTPUT_DIR}')
    print('篩選條件: 無人臉 / 無開箱 / 每品最少 3 支合併')
    print('=' * 55)

    os.makedirs(OUTPUT_DIR, exist_ok=True)

    wb = openpyxl.load_workbook(EXCEL_PATH)
    ws = wb.active
    total = ws.max_row - 1
    done_count = 0

    async with async_playwright() as p:
        print('\n連接 Chrome (port 9222)...')
        browser = await p.chromium.connect_over_cdp('http://127.0.0.1:9222')
        ctx = browser.contexts[0]

        pg = next((pg for pg in ctx.pages
                   if pg.url.startswith('https://shopee.tw') and 'affiliate' not in pg.url), None)
        if not pg:
            pg = await ctx.new_page()
            await pg.goto('https://shopee.tw', wait_until='domcontentloaded', timeout=20000)
            await asyncio.sleep(3)
        print(f'頁面: {pg.url[:60]}')

        for row in ws.iter_rows(min_row=2):
            name        = row[COL_NAME   - 1].value
            link        = row[COL_LINK   - 1].value
            status_cell = row[COL_STATUS - 1]
            row_idx     = row[0].row - 1   # 1-based product index

            if not name or not link: continue
            if status_cell.value in ('影片完成', '無評論影片'):
                print(f'[skip {row_idx:3}] {str(name)[:25]} ({status_cell.value})')
                continue

            print(f'\n[{row_idx:3}/{total}] {str(name)[:40]}')

            status = await process_product(pg, name, link, row_idx)
            status_cell.value = status
            wb.save(EXCEL_PATH)

            if status == '影片完成': done_count += 1

            await asyncio.sleep(random.uniform(1.5, 3.0))

        pass  # keep browser open

    print(f'\n{"="*55}')
    print(f'完成！成功製作 {done_count} 個商品影片')
    print(f'影片存放：{OUTPUT_DIR}')

asyncio.run(main())
