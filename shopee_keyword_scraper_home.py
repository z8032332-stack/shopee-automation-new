# 蝦皮關鍵字選品主程式
# CDP 連線至 Chrome port 9222，所有設定從 .env 讀取
import sys, io, asyncio, json, os, random
sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8', errors='replace')
sys.stderr = io.TextIOWrapper(sys.stderr.buffer, encoding='utf-8', errors='replace')

from dotenv import load_dotenv
load_dotenv(dotenv_path=os.path.join(os.path.dirname(os.path.abspath(__file__)), '.env'))

from playwright.async_api import async_playwright
from urllib.parse import quote
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

ALL_KEYWORDS = [
    # 3C 周邊
    '手機殼','行動電源','藍芽耳機','耳機','藍芽喇叭',
    '充電線','無線充電','保護貼','手機支架',
    # 包包/行李
    '行李箱','後背包','側背包','手提包','腰包',
    '托特包','購物袋','學生書包',
    # 穿搭
    '外套','連身裙','洋裝','牛仔褲','T恤',
    '運動鞋','涼鞋','拖鞋','靴子','帆布鞋',
    '防曬衣','運動服','泳衣',
    # 飾品
    '耳環','項鍊','手鍊','戒指','髮夾','髮圈',
    # 居家生活
    '保溫杯','水壺','便當盒','收納箱','掛勾',
    '除濕機','電風扇','延長線','LED燈','香氛',
    # 寢具
    '棉被','枕頭','床墊','寢具組',
    # 廚房
    '氣炸鍋','保鮮盒','餐具組','鍋具',
    # 美妝保養
    '防曬乳','面膜','精華液','口紅','眉筆',
    # 玩具/兒童
    '兒童玩具','積木','玩具車','娃娃','存錢筒',
    # 寵物
    '貓零食','狗零食','寵物玩具','貓砂',
    # 運動
    '瑜伽墊','跳繩','啞鈴','運動手套',
    # 季節熱搜
    '雨傘','安全帽','除蟲燈','暖暖包',
]
BLACKLIST  = ['藥','酒','菸','棉花棒','化妝棉','直播','下單','1元','專拍','鏈接']
TARGET     = int(os.getenv('KEYWORD_TARGET',    '1000'))
MIN_SALES  = int(os.getenv('KEYWORD_MIN_SALES', '100'))
MAX_PER_KW = (TARGET // len(ALL_KEYWORDS)) + 5
OUTPUT     = os.getenv('KEYWORD_OUTPUT', r'D:\Users\user\Desktop\蝦皮影片專案\蝦皮選品_2026年5月.xlsx')

PRODUCT_HISTORY_FILE = os.path.join(os.path.dirname(os.path.abspath(__file__)), 'product_history.json')

def load_product_history():
    if os.path.exists(PRODUCT_HISTORY_FILE):
        with open(PRODUCT_HISTORY_FILE, encoding='utf-8') as f:
            return json.load(f)
    return {}

def save_product_history(new_products):
    from datetime import date
    h = load_product_history()
    today = date.today().isoformat()
    for p in new_products:
        uid = f"{p['shop_id']}_{p['item_id']}"
        h[uid] = today
    with open(PRODUCT_HISTORY_FILE, 'w', encoding='utf-8') as f:
        json.dump(h, f, ensure_ascii=False, indent=2)
    print(f'  商品歷史已更新（累計 {len(h)} 筆）：{PRODUCT_HISTORY_FILE}')

KEYWORDS = ALL_KEYWORDS[:]
random.shuffle(KEYWORDS)

def is_bl(name): return any(k in name for k in BLACKLIST)
def pp(raw):
    try:
        v = int(raw)
        return str(v // 100000) if v > 100000 else str(v)
    except: return str(raw)

async def safe_eval(page, js, retries=3):
    for i in range(retries):
        try: return await page.evaluate(js)
        except:
            if i < retries - 1: await asyncio.sleep(1.5)
    return None

async def search_kw(page, kw, limit=50, page_offset=0):
    enc = quote(kw)
    js = """
    (async () => {
      try {
        const r = await fetch('https://affiliate.shopee.tw/api/v3/offer/product/list?list_type=0&sort_type=1&page_offset=PAGE&page_limit=LIMIT&keyword=KW', {credentials:'include'});
        const d = await r.json();
        return d;
      } catch(e) { return {error: e.toString()}; }
    })()
    """.replace('LIMIT', str(limit)).replace('KW', enc).replace('PAGE', str(page_offset))
    d = await safe_eval(page, js) or {}
    items = []
    for it in (d.get('data') or {}).get('list') or []:
        info = it.get('batch_item_for_item_card_full') or {}
        name = info.get('name', '')
        if is_bl(name): continue
        sales = info.get('sold') or info.get('historical_sold') or 0
        if sales < MIN_SALES: continue
        price = pp(info.get('price') or info.get('price_min', 0))
        shop_id = str(info.get('shopid', ''))
        item_id = str(info.get('itemid', '') or it.get('item_id', ''))
        items.append({
            'name': name, 'price': price, 'sales': sales, 'keyword': kw,
            'comm_rate': it.get('default_commission_rate', '') or it.get('seller_commission_rate', ''),
            'affiliate_link': it.get('long_link', ''),
            'product_url': it.get('product_link', ''),
            'shop_id': shop_id, 'item_id': item_id,
        })
    return items

async def get_link(page, url):
    if not url: return ''
    enc = quote(url, safe='')
    js = """
    (async () => {
      try {
        const r = await fetch('https://affiliate.shopee.tw/api/v3/product/get_affiliate_link?product_url=URL', {credentials:'include'});
        const d = await r.json();
        return (d && d.data && (d.data.short_link || d.data.link)) || '';
      } catch(e) { return ''; }
    })()
    """.replace('URL', enc)
    return await safe_eval(page, js) or ''

def build_excel(products, append=False):
    thin = Side(style='thin', color='DDDDDD')
    bd = Border(left=thin, right=thin, top=thin, bottom=thin)
    hdrs = ['編號','品名','分潤連結','價格','分潤率','銷量','對應關鍵字','文案','標題','狀態']
    wds  = [8, 55, 48, 10, 10, 10, 16, 30, 30, 12]

    if append and os.path.exists(OUTPUT):
        wb = openpyxl.load_workbook(OUTPUT)
        ws = wb.active
        start_row = ws.max_row + 1
        start_no  = ws.max_row
        print(f'  Append 模式：從第 {start_row} 行接續（編號從 {start_no} 開始）')
    else:
        wb = openpyxl.Workbook(); ws = wb.active; ws.title = '蝦皮關鍵字選品'
        for ci, (h, w) in enumerate(zip(hdrs, wds), 1):
            c = ws.cell(1, ci, h)
            c.font = Font(name='微軟正黑體', bold=True, color='FFFFFF', size=11)
            c.fill = PatternFill('solid', fgColor='C0392B')
            c.border = bd
            c.alignment = Alignment(horizontal='center', vertical='center')
            ws.column_dimensions[get_column_letter(ci)].width = w
        ws.row_dimensions[1].height = 28
        start_row = 2; start_no = 1

    for i, p in enumerate(products):
        ri = start_row + i
        no = start_no + i
        fill = PatternFill('solid', fgColor='FFF5F5' if ri % 2 == 0 else 'FFFFFF')
        vals = [no, p.get('name',''), p.get('affiliate_link',''), f"${p.get('price','0')}",
                p.get('comm_rate',''), p.get('sales',0), p.get('keyword',''), '', '', '']
        for ci, val in enumerate(vals, 1):
            c = ws.cell(ri, ci, val); c.fill = fill; c.border = bd
            c.font = Font(name='Arial', size=10, color='0563C1', underline='single') if ci == 3 \
                     else Font(name='微軟正黑體', size=10)
            c.alignment = Alignment(horizontal='left' if ci in (2, 3) else 'center',
                                    vertical='center', wrap_text=(ci == 2))
        ws.row_dimensions[ri].height = 32

    ws.freeze_panes = 'A2'
    ws.auto_filter.ref = f'A1:J{ws.max_row}'
    wb.save(OUTPUT)
    print(f'Excel 已儲存：{OUTPUT}（共 {ws.max_row - 1} 筆）')

async def main():
    print('=' * 50)
    print(f'蝦皮關鍵字選品（目標 {TARGET} 筆，銷量 >= {MIN_SALES}）')
    print('=' * 50)

    async with async_playwright() as p:
        print('\n連接 Chrome (port 9222)...')
        browser = await p.chromium.connect_over_cdp('http://127.0.0.1:9222')
        ctx = browser.contexts[0]

        pg_aff = next((pg for pg in ctx.pages if 'affiliate.shopee.tw' in pg.url), None)
        if not pg_aff:
            pg_aff = await ctx.new_page()
            await pg_aff.goto('https://affiliate.shopee.tw/offer/product_offer',
                              wait_until='domcontentloaded', timeout=20000)
            await asyncio.sleep(3)
        print(f'affiliate 頁面：{pg_aff.url}')

        test = await search_kw(pg_aff, '手機殼', limit=2)
        print(f'API 測試：{"OK - " + str(len(test)) + "筆" if test else "FAIL"}')

        prev_uids = set(load_product_history().keys())
        print(f'歷史商品 {len(prev_uids)} 筆（本次將排除）')

        # 搜尋（每個關鍵字最多 5 頁）
        print(f'\n搜尋關鍵字（每關鍵字上限 {MAX_PER_KW} 筆，最多 5 頁）...')
        all_p = []; seen = set(); kw_count = {}
        for kw in KEYWORDS:
            if len(all_p) >= TARGET * 2: break
            kw_total = 0
            for page in range(5):
                if kw_count.get(kw, 0) >= MAX_PER_KW: break
                items = await search_kw(pg_aff, kw, limit=50, page_offset=page)
                if not items: break
                added = 0
                for item in items:
                    if kw_count.get(kw, 0) >= MAX_PER_KW: break
                    uid = f"{item['shop_id']}_{item['item_id']}"
                    if uid in seen or uid == '_': continue
                    if uid in prev_uids: continue
                    seen.add(uid); all_p.append(item)
                    kw_count[kw] = kw_count.get(kw, 0) + 1
                    added += 1
                kw_total += added
                if added == 0: break
                await asyncio.sleep(random.uniform(0.8, 1.2))
            print(f'  [{kw}] {kw_total}筆 累計{len(all_p)}')
            await asyncio.sleep(random.uniform(0.5, 1.0))

        # 驗證連結
        print(f'\n候選 {len(all_p)} 筆，驗證分潤連結...')
        valid = []; nl = 0
        for item in all_p:
            if len(valid) >= TARGET: break
            lnk = item.get('affiliate_link', '')
            if not lnk:
                lnk = await get_link(pg_aff, item['product_url'])
                item['affiliate_link'] = lnk
            if not lnk:
                nl += 1; continue
            valid.append(item)
        print(f'有效 {len(valid)} 筆 | 無連結 {nl} 筆')

    if len(valid) == 0:
        print('⚠️  0 筆有效商品，不輸出 Excel。')
        return

    append_mode = os.getenv('KEYWORD_APPEND', '0') == '1'
    build_excel(valid[:TARGET], append=append_mode)
    save_product_history(valid[:TARGET])
    print(f'\n完成！共 {len(valid[:TARGET])} 筆')

asyncio.run(main())
