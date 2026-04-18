"""
gen_copy.py — 用 Gemini 批量生成文案+標題，填入 Excel H/I 欄
"""
import sys, io, os, time
sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8', errors='replace')
sys.stderr = io.TextIOWrapper(sys.stderr.buffer, encoding='utf-8', errors='replace')

from dotenv import load_dotenv
load_dotenv(dotenv_path=os.path.join(os.path.dirname(os.path.abspath(__file__)), '.env'))

import openpyxl
from groq import Groq

EXCEL_PATH = os.getenv('EXCEL_PATH', r'C:\Users\user\Desktop\蝦皮素材\蝦皮選品_2026年 (1).xlsx')
GROQ_KEY   = os.getenv('GROQ_KEY', '')

COL_NAME   = 2   # B
COL_LINK   = 3   # C
COL_COPY   = 8   # H
COL_TITLE  = 9   # I
COL_STATUS = 10  # J

client = Groq(api_key=GROQ_KEY)

COPY_PROMPT = """\
你是蝦皮短影音文案專家。根據以下商品名稱，生成台灣口語化的口播文案和標題。

商品名稱：{name}

【文案規則】
- 共10句，每句一行，不加編號
- 每句約10個中文字，最多15字，絕對不能超過
- 第1-2句：開場鉤子（提出問題或痛點）
- 第3-4句：說明商品主題
- 第5-6句：鋪陳問題讓觀眾有感
- 第7-8句：解決方案（商品特點）
- 第9句：成果想像
- 第10句：購買CTA（每次不同，要有創意，用「下方連結」）
- 語氣：活潑輕鬆，像朋友推薦
- 每句結尾加句號
- 禁用詞：治療、保證、消炎、藥效、永久

【標題規則】
- 1行，不可有任何表情符號或特殊符號
- 加5-8個 #hashtag（中文關鍵字）
- 結尾加1句短的商品賣點說明
- 全部合計不超過140字

【回答格式】（只輸出內容，不要其他說明）
文案：
（10句文案）
標題：
（標題內容）
"""

def generate(name):
    prompt = COPY_PROMPT.format(name=name.strip())
    for attempt in range(3):
        try:
            resp = client.chat.completions.create(
                model='llama-3.3-70b-versatile',
                messages=[{'role': 'user', 'content': prompt}],
                temperature=0.8,
            )
            text = resp.choices[0].message.content.strip()
            # 解析
            copy_part, title_part = '', ''
            if '文案：' in text and '標題：' in text:
                parts = text.split('標題：')
                copy_raw = parts[0].replace('文案：', '').strip()
                title_part = parts[1].strip()
                # 清理文案：移除空行、編號
                lines = [l.strip() for l in copy_raw.splitlines() if l.strip()]
                copy_part = '\n'.join(lines)
                # 標題：移除 emoji/特殊符號，截斷至140字
                import re as _re
                title_part = _re.sub(r'[\U00010000-\U0010FFFF]', '', title_part)  # 移除 emoji
                title_part = _re.sub(r'[✨💰🔊🎵🪙✅❌🎯🏆💡🛒🔥⭐️]', '', title_part)  # 常見 emoji
                title_part = title_part.strip()[:140]
            return copy_part, title_part
        except Exception as e:
            print(f'  [retry {attempt+1}] {e}')
            time.sleep(3)
    return '', ''

def main():
    wb = openpyxl.load_workbook(EXCEL_PATH)
    ws = wb.active
    total = ws.max_row - 1
    done = 0

    for row in ws.iter_rows(min_row=2):
        name        = row[COL_NAME - 1].value
        link        = row[COL_LINK - 1].value
        copy_cell   = row[COL_COPY - 1]
        title_cell  = row[COL_TITLE - 1]
        row_idx     = row[0].row

        if not name or not link: continue
        if not str(link).startswith('http'): continue
        if copy_cell.value:  # 已有文案就跳過
            print(f'[skip {row_idx}] 已有文案')
            continue

        print(f'\n[{row_idx:>3}/{total}] {str(name)[:40]}')
        copy_text, title_text = generate(str(name))

        if copy_text:
            copy_cell.value  = copy_text
            title_cell.value = title_text
            done += 1
            print(f'  ✓ 文案 {len(copy_text)}字 / 標題 {len(title_text)}字')
        else:
            print(f'  ✗ 生成失敗')

        wb.save(EXCEL_PATH)
        time.sleep(1.5)  # 避免 API rate limit

    print(f'\n完成！共生成 {done} 筆文案')
    print(f'存檔：{EXCEL_PATH}')

if __name__ == '__main__':
    main()
