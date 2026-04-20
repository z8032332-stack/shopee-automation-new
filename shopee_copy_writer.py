"""
shopee_copy_writer.py
─────────────────────────────────────────────────
用 Gemini 自動生成蝦皮短影音文案 + 標題，寫回 Excel
- 每分鐘最多 5 次 API 請求（每次間隔 13 秒）
- 文案 + 標題合併一次請求（省 quota）
- 跳過已有文案 or 狀態欄含「影片完成」/「直播」的列
- 每寫完一筆立即存檔（防崩潰遺失）
─────────────────────────────────────────────────
用法：
  python shopee_copy_writer.py
  python shopee_copy_writer.py --start 50    # 從第 50 筆開始（覆蓋 .env START_ROW）
  python shopee_copy_writer.py --overwrite   # 強制覆寫已有文案的列
"""

import os
import sys
import time
import argparse
import openpyxl
from dotenv import load_dotenv

# ── 強制 stdout/stderr 用 UTF-8（避免 cp950 無法印出 emoji 或中文）──
if hasattr(sys.stdout, 'reconfigure'):
    sys.stdout.reconfigure(encoding='utf-8', errors='replace')
if hasattr(sys.stderr, 'reconfigure'):
    sys.stderr.reconfigure(encoding='utf-8', errors='replace')

# ── 載入設定 ──────────────────────────────────
load_dotenv()

EXCEL_PATH = os.getenv('EXCEL_PATH', '')
GEMINI_KEY = os.getenv('GEMINI_KEY', '')
GROQ_KEY   = os.getenv('GROQ_KEY', '')
START_ROW  = int(os.getenv('START_ROW', 2))
CLIPS_DIR  = os.getenv('OUTPUT_DIR', '')   # output_videos/_clips_XXX 所在資料夾

COL_NAME   = int(os.getenv('COL_NAME',   2))
COL_LINK   = int(os.getenv('COL_LINK',   3))
COL_COPY   = int(os.getenv('COL_COPY',   8))
COL_TITLE  = int(os.getenv('COL_TITLE',  9))
COL_STATUS = int(os.getenv('COL_STATUS', 10))

# ── Rate limit ────────────────────────────────
# Groq 免費版每分鐘 30 次，間隔 3 秒即可
DELAY_SEC = 3
GROQ_MODEL = 'llama-3.1-8b-instant'
GEMINI_MODELS = ['gemini-2.0-flash-lite', 'gemini-2.0-flash']

# 延遲初始化，避免模組載入時卡住
_groq_client = None
_gemini_client = None

def _get_groq():
    global _groq_client
    if _groq_client is None:
        from groq import Groq
        _groq_client = Groq(api_key=GROQ_KEY)
    return _groq_client

def _get_gemini():
    global _gemini_client
    if _gemini_client is None and GEMINI_KEY:
        from google import genai as _genai
        _gemini_client = _genai.Client(api_key=GEMINI_KEY)
    return _gemini_client


def build_prompt(product_name: str, link: str) -> str:
    return f"""請針對以下商品，依照規則分別生成【文案】和【標題】，不要加任何其他說明。

商品名稱：{product_name}

【文案】
規則：
- 生成一段 20~30 秒的口播文案，繁體中文
- 每句話不超過 10 個字
- 遇到標點符號就換行，每行最後都改為句號
- 一句話一行，總共 10 行
- 結構：
  第1~2行：開場鉤子（0-5秒）
  第3~4行：主題說明（6-10秒）
  第5~6行：問題鋪成（11-15秒）
  第7~8行：解決辦法（16-20秒）
  第9行：成果想像（21-25秒）
  第10行：CTA互動引導（結尾）
- 禁止用「你是否」「你有沒有」「你是不是」等問句開頭

【標題】
規則：
- 先寫 1~2 句話，說明消費者買了這個商品有什麼好處（放最前面）
- 把商品名稱中每個關鍵詞前面加 # 字鍵
- 再加入高相關度的關鍵詞，前面也加 # 字鍵
- 全部不換行，連成一段
- 避開品牌詞與店名
- 不要使用任何表情符號或特殊符號
- 總共湊滿 140 字，不夠的字數用 #關鍵字 補足"""


def parse_response(text: str):
    """從 Gemini 回應中解析文案與標題"""
    copy  = ''
    title = ''

    if '【文案】' in text and '【標題】' in text:
        parts = text.split('【標題】')
        copy_part  = parts[0].replace('【文案】', '').strip()
        title_part = parts[1].strip() if len(parts) > 1 else ''
        copy  = copy_part.strip()
        title = title_part.strip()
    else:
        # 無法解析格式 → 整段當作文案
        copy = text.strip()

    return copy, title


def generate(product_name: str, link: str):
    """先用 Groq，失敗或額度滿再換 Gemini"""
    prompt = build_prompt(product_name, link)

    # ── 1. Groq（主力，每天 14,400 次免費） ──────
    try:
        resp = _get_groq().chat.completions.create(
            model=GROQ_MODEL,
            messages=[{'role': 'user', 'content': prompt}],
            temperature=0.7,
            max_tokens=512,
        )
        text = resp.choices[0].message.content
        return parse_response(text), f'groq/{GROQ_MODEL}'
    except Exception as e:
        err_str = str(e)
        if '429' in err_str or 'rate' in err_str.lower():
            print(f'    ⚠ Groq 額度/速率限制，換 Gemini...')
        else:
            print(f'    ⚠ Groq 錯誤（{err_str[:60]}），換 Gemini...')

    # ── 2. Gemini 備用 ─────────────────────────
    gc = _get_gemini()
    if gc:
        for model in GEMINI_MODELS:
            try:
                resp = gc.models.generate_content(model=model, contents=prompt)
                return parse_response(resp.text), f'gemini/{model}'
            except Exception as e:
                err_str = str(e)
                if 'RESOURCE_EXHAUSTED' in err_str or '429' in err_str:
                    print(f'    ⚠ {model} 額度已滿，換下一個...')
                    continue
                print(f'    ⚠ {model} 錯誤（{err_str[:60]}），跳過...')
                continue

    raise Exception(
        '所有模型每日額度已用完！\n'
        '解決方式：\n'
        '  1. 等明天 UTC 00:00 額度重置後再跑\n'
        '  2. Groq 額度：console.groq.com 查看用量'
    )


def is_yellow_row(row):
    """偵測品名欄（col B）是否被標記黃色（FFFFFF00），黃色 = 直播/垃圾，要跳過"""
    cell = row[COL_NAME - 1]  # 只看品名欄
    fill = cell.fill
    if fill and fill.fgColor and fill.fgColor.type == 'rgb':
        return fill.fgColor.rgb == 'FFFFFF00'
    return False


def main():
    parser = argparse.ArgumentParser()
    parser.add_argument('--start',          type=int, default=None, help='從第幾筆開始（Excel row，含標題行）')
    parser.add_argument('--overwrite',      action='store_true',    help='強制覆寫已有文案+標題')
    parser.add_argument('--overwrite-title',action='store_true',    help='只覆寫標題，不動文案')
    args = parser.parse_args()

    if not EXCEL_PATH or not os.path.exists(EXCEL_PATH):
        print(f'[錯誤] 找不到 Excel：{EXCEL_PATH}')
        print('請確認 .env 的 EXCEL_PATH 設定正確。')
        sys.exit(1)

    if not GEMINI_KEY:
        print('[錯誤] .env 缺少 GEMINI_KEY')
        sys.exit(1)

    start_row = args.start if args.start else START_ROW

    print(f'[設定] Excel：{EXCEL_PATH}')
    print(f'[設定] 起始列：{start_row}，覆寫模式：{args.overwrite}')
    print(f'[設定] 速率限制：每 {DELAY_SEC} 秒 1 次（每分鐘最多 5 次）')
    print('─' * 55)

    # ── 開檔前先備份，防止 Ctrl+C 中途損毀 ──────
    import shutil
    backup_path = EXCEL_PATH.replace('.xlsx', '_backup.xlsx')
    shutil.copy2(EXCEL_PATH, backup_path)

    wb = openpyxl.load_workbook(EXCEL_PATH)
    ws = wb.active

    # 先掃一遍算出待處理數量
    pending = []
    for i, row in enumerate(ws.iter_rows(min_row=start_row), start=start_row):
        name   = row[COL_NAME   - 1].value
        copy   = row[COL_COPY   - 1].value
        status = row[COL_STATUS - 1].value
        if not name:
            continue
        if is_yellow_row(row):          # 跳過黃色（直播/垃圾）列
            continue
        if status == '影片完成':
            continue
        if status and '直播' in str(status):
            continue
        if status == 'no_ids':
            continue
        title  = row[COL_TITLE  - 1].value
        # --overwrite-title：只要有文案就進來（不管有沒有標題）
        # 一般模式：有文案又有標題才跳過
        if args.overwrite_title:
            if not copy:
                continue  # 沒文案也沒標題，整列跳過
        elif copy and title and not args.overwrite:
            continue
        # 跳過沒有影片的列（狀態欄沒有 clips_ok 代表還沒抓影片）
        if not status or 'clips_ok' not in str(status):
            continue
        pending.append(i)

    total = len(pending)
    if total == 0:
        print('沒有需要生成的列（全部已有文案或已完成）。')
        print('提示：加 --overwrite 可強制重新生成。')
        return

    eta_min = total * DELAY_SEC / 60
    print(f'待生成：{total} 筆，預估約 {eta_min:.1f} 分鐘')
    print('─' * 55)

    done = 0
    fail = 0

    for seq, row_idx in enumerate(pending, 1):
        row = list(ws.iter_rows(min_row=row_idx, max_row=row_idx))[0]

        name  = row[COL_NAME  - 1].value
        link  = row[COL_LINK  - 1].value or ''
        title_cell = row[COL_TITLE - 1]
        copy_cell  = row[COL_COPY  - 1]

        print(f'[{seq}/{total}] {str(name)[:40]}...')

        try:
            (copy, title), used_model = generate(name, link)

            if not args.overwrite_title:                # --overwrite-title 時不動文案
                if not copy_cell.value or args.overwrite:
                    copy_cell.value = copy
            if title:                                   # 標題一律覆寫
                title_cell.value = title

            wb.save(EXCEL_PATH)
            done += 1
            print(f'    OK [{used_model}] 文案 {len(copy)} 字 | 標題：{title[:25] if title else "（已有）"}')

        except Exception as e:
            fail += 1
            print(f'    NG 失敗，跳過：{e}')
            if '所有模型每日額度已用完' in str(e):
                print('\n程式終止：等明天再跑，或換新的 API Key。')
                break

        # 最後一筆不需要等
        if seq < total:
            remaining = (total - seq) * DELAY_SEC
            print(f'    wait {DELAY_SEC}s... (剩餘約 {remaining // 60} 分 {remaining % 60} 秒)')
            time.sleep(DELAY_SEC)

    print('─' * 55)
    print(f'完成！成功 {done} 筆，失敗 {fail} 筆，Excel 已存檔。')


if __name__ == '__main__':
    main()
